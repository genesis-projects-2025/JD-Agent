from fastapi import APIRouter, Depends, HTTPException, File, Form, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, text
from pydantic import BaseModel
from typing import Optional

from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from jose import jwt, JWTError

from app.core.database import get_db
from app.core.config import settings
from app.core.security import create_access_token
from app.models.jd_session_model import JDSession
from app.models.user_model import Employee
from app.services.kra_kpi_service import process_kra_kpi_document
import logging

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Admin"])
security = HTTPBearer()


class AdminLoginRequest(BaseModel):
    code: str
    password: str


class AdminLoginResponse(BaseModel):
    token: str
    role: str


class StatCardData(BaseModel):
    total_employees: int
    total_generated_jds: int
    pending_jds: int
    approved_jds: int
    rejected_jds: int


async def get_current_admin(
    credentials: HTTPAuthorizationCredentials = Depends(security),
):
    """Dependency to protect routes — verifies the JWT token."""
    token = credentials.credentials
    try:
        payload = jwt.decode(
            token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM]
        )
        role = payload.get("sub")
        if role != "ADMIN":
            raise HTTPException(status_code=403, detail="Not authorized as admin")
        return role
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired admin token")


@router.post("/auth/admin-login", response_model=AdminLoginResponse)
async def admin_login(request: AdminLoginRequest):
    # Use config-backed credentials with space trimming and case insensitivity on code
    code_input = (request.code or "").strip()
    pass_input = (request.password or "").strip()
    
    if (
        code_input.lower() == settings.ADMIN_CODE.lower()
        and pass_input == settings.ADMIN_PASSWORD
    ):
        token = create_access_token(subject="ADMIN")
        return AdminLoginResponse(token=token, role="ADMIN")

    raise HTTPException(status_code=401, detail="Invalid admin credentials")


_ADMIN_STATS_CACHE: dict = {}
_ADMIN_CHARTS_CACHE: dict = {}
_ADMIN_CACHE_TTL = 30.0


@router.get("/admin/stats/overview", response_model=StatCardData)
async def get_admin_overview(
    db: AsyncSession = Depends(get_db), admin_role: str = Depends(get_current_admin)
):
    import time
    now = time.time()
    if _ADMIN_STATS_CACHE and (now - _ADMIN_STATS_CACHE.get("ts", 0)) < _ADMIN_CACHE_TTL:
        return _ADMIN_STATS_CACHE["data"]

    # Single-pass SQL query for accurate overview stats using latest per-employee JD session
    res = await db.execute(
        text("""
        WITH RankedJDs AS (
            SELECT 
                employee_id, status,
                ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
            FROM jd_sessions
        ),
        LatestJDs AS (
            SELECT employee_id, status
            FROM RankedJDs WHERE rn = 1
        )
        SELECT
            (SELECT COUNT(*) FROM organogram) as total_employees,
            (SELECT COUNT(DISTINCT employee_id) FROM LatestJDs WHERE status IN ('jd_generated', 'ready_for_generation', 'sent_to_manager', 'sent_to_hr', 'approved', 'manager_rejected', 'hr_rejected', 'rejected')) as total_generated_jds,
            (SELECT COUNT(DISTINCT employee_id) FROM LatestJDs WHERE status IN ('sent_to_manager', 'sent_to_hr')) as pending_jds,
            (SELECT COUNT(DISTINCT employee_id) FROM LatestJDs WHERE status = 'approved') as approved_jds,
            (SELECT COUNT(DISTINCT employee_id) FROM LatestJDs WHERE status IN ('manager_rejected', 'hr_rejected', 'rejected')) as rejected_jds
    """)
    )
    row = res.mappings().first() or {}
    result_data = StatCardData(
        total_employees=row.get("total_employees", 0),
        total_generated_jds=row.get("total_generated_jds", 0),
        pending_jds=row.get("pending_jds", 0),
        approved_jds=row.get("approved_jds", 0),
        rejected_jds=row.get("rejected_jds", 0),
    )
    _ADMIN_STATS_CACHE["data"] = result_data
    _ADMIN_STATS_CACHE["ts"] = now
    return result_data


@router.get("/admin/stats/charts")
async def get_admin_charts(
    db: AsyncSession = Depends(get_db), admin_active: str = Depends(get_current_admin)
):
    import time
    now = time.time()
    if _ADMIN_CHARTS_CACHE and (now - _ADMIN_CHARTS_CACHE.get("ts", 0)) < _ADMIN_CACHE_TTL:
        return _ADMIN_CHARTS_CACHE["data"]

    # 1. Pipeline Chart (Bar Chart) - based on latest JD status per employee
    pipeline_res = await db.execute(
        text("""
            WITH RankedJDs AS (
                SELECT 
                    status,
                    ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
                FROM jd_sessions
            )
            SELECT status, COUNT(*) as count
            FROM RankedJDs WHERE rn = 1
            GROUP BY status
        """)
    )
    status_counts = pipeline_res.fetchall()

    pipeline_data = [{"status": row[0], "count": row[1]} for row in status_counts]

    pipeline_map = {item["status"]: item["count"] for item in pipeline_data}
    normalized_pipeline = [
        {
            "status": "Drafting",
            "count": pipeline_map.get("collecting", 0)
            + pipeline_map.get("draft", 0)
            + pipeline_map.get("jd_generated", 0)
            + pipeline_map.get("ready_for_generation", 0),
        },
        {"status": "Pending Manager", "count": pipeline_map.get("sent_to_manager", 0)},
        {"status": "Pending HR", "count": pipeline_map.get("sent_to_hr", 0)},
        {"status": "Approved", "count": pipeline_map.get("approved", 0)},
        {
            "status": "Rejected",
            "count": pipeline_map.get("manager_rejected", 0)
            + pipeline_map.get("hr_rejected", 0)
            + pipeline_map.get("rejected", 0),
        },
    ]

    # 2. Manager Response Chart (Doughnut)
    manager_responded = (
        pipeline_map.get("sent_to_hr", 0)
        + pipeline_map.get("manager_rejected", 0)
        + pipeline_map.get("hr_rejected", 0)
        + pipeline_map.get("approved", 0)
    )
    manager_pending = pipeline_map.get("sent_to_manager", 0)

    response_rate = [
        {"name": "Responded", "value": manager_responded},
        {"name": "Pending", "value": manager_pending},
    ]

    res_dict = {"pipeline": normalized_pipeline, "manager_response": response_rate}
    _ADMIN_CHARTS_CACHE["data"] = res_dict
    _ADMIN_CHARTS_CACHE["ts"] = now
    return res_dict


@router.get("/admin/users")
async def get_admin_users(
    role: Optional[str] = None,
    status: Optional[str] = None,
    department: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 1000,
    db: AsyncSession = Depends(get_db),
    admin_active: str = Depends(get_current_admin),
):
    sql = """
        WITH RankedJDs AS (
            SELECT 
                id, employee_id, status, updated_at,
                ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
            FROM jd_sessions
        ),
        LatestJDs AS (
            SELECT id, employee_id, status, updated_at
            FROM RankedJDs WHERE rn = 1
        ),
        RankedKRAs AS (
            SELECT 
                id, employee_id, status, updated_at,
                ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
            FROM kra_kpi_sessions
        ),
        LatestKRAs AS (
            SELECT id, employee_id, status, updated_at
            FROM RankedKRAs WHERE rn = 1
        )
        SELECT 
            o.code as employee_id,
            o.employee_name as name,
            e.email as email,
            o.department as department,
            o.designation as role,
            o.reporting_manager as manager_name,
            COALESCE(js.status, 'No JD') as jd_status,
            js.id::text as jd_session_id,
            js.updated_at as last_active,
            CASE 
                WHEN uk.id IS NOT NULL THEN 'approved'
                ELSE COALESCE(ks.status, 'Not Started')
            END as kra_kpi_status
        FROM organogram o
        LEFT JOIN employees e ON e.id = o.code
        LEFT JOIN LatestJDs js ON js.employee_id = o.code
        LEFT JOIN LatestKRAs ks ON ks.employee_id = o.code
        LEFT JOIN uploaded_kra_kpis uk ON uk.employee_id = o.code
        WHERE 1=1
    """
    params = {}
    if role:
        sql += " AND o.designation ILIKE :role"
        params["role"] = f"%{role}%"
    if department and department.strip() and department.lower() != "all":
        sql += " AND LOWER(TRIM(o.department)) = LOWER(:department)"
        params["department"] = department.strip()
    if status:
        if status.lower() == "no jd":
            sql += " AND js.id IS NULL"
        else:
            sql += " AND js.status = :status"
            params["status"] = status
    if search:
        sql += " AND (o.employee_name ILIKE :search OR o.code ILIKE :search OR e.email ILIKE :search OR o.department ILIKE :search)"
        params["search"] = f"%{search}%"

    sql += " ORDER BY o.employee_name ASC"
    
    # Cap limit to prevent OOM while allowing full company list
    safe_limit = max(1, min(limit, 2000))
    safe_skip = max(0, skip)
    
    sql += " LIMIT :limit OFFSET :skip"
    params["limit"] = safe_limit
    params["skip"] = safe_skip

    result = await db.execute(text(sql), params)
    rows = result.mappings().all()

    formatted_results = []
    seen = set()
    for r in rows:
        emp_id = r["employee_id"]
        if emp_id in seen:
            continue
        seen.add(emp_id)
        formatted_results.append({
            "employee_id": emp_id,
            "name": r["name"] or "Unknown",
            "email": r["email"],
            "department": r["department"],
            "role": r["role"] or "Employee",
            "manager_name": r["manager_name"],
            "jd_status": r["jd_status"],
            "jd_session_id": r["jd_session_id"],
            "kra_kpi_status": r["kra_kpi_status"],
            "last_active": r["last_active"].isoformat() if r.get("last_active") else None,
        })

    return {
        "items": formatted_results,
        "skip": safe_skip,
        "limit": safe_limit,
        "count": len(formatted_results),
    }


@router.get("/admin/departments/summary")
async def get_department_summary(
    db: AsyncSession = Depends(get_db),
    admin_active: str = Depends(get_current_admin),
):
    """
    Returns department-wise JD & KRA/KPI completion summary across the company using latest sessions per employee.
    """
    sql = """
        WITH RankedJDs AS (
            SELECT 
                id, employee_id, status, updated_at,
                ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
            FROM jd_sessions
        ),
        LatestJDs AS (
            SELECT id, employee_id, status, updated_at
            FROM RankedJDs WHERE rn = 1
        ),
        RankedKRAs AS (
            SELECT 
                id, employee_id, status, updated_at,
                ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
            FROM kra_kpi_sessions
        ),
        LatestKRAs AS (
            SELECT id, employee_id, status, updated_at
            FROM RankedKRAs WHERE rn = 1
        )
        SELECT 
            COALESCE(NULLIF(TRIM(o.department), ''), 'Unassigned') as department,
            COUNT(DISTINCT o.code) as total_employees,
            COUNT(DISTINCT CASE WHEN js.status = 'approved' THEN o.code END) as jd_completed,
            COUNT(DISTINCT CASE WHEN js.status = 'sent_to_manager' THEN o.code END) as pending_manager,
            COUNT(DISTINCT CASE WHEN js.status = 'sent_to_hr' THEN o.code END) as pending_hr,
            COUNT(DISTINCT CASE WHEN js.status IN ('collecting', 'draft', 'jd_generated', 'ready_for_generation', 'manager_rejected', 'hr_rejected') THEN o.code END) as in_progress,
            COUNT(DISTINCT CASE WHEN js.id IS NULL THEN o.code END) as not_started,
            COUNT(DISTINCT CASE WHEN uk.id IS NOT NULL OR ks.status IN ('confirmed', 'approved') THEN o.code END) as kra_completed
        FROM organogram o
        LEFT JOIN LatestJDs js ON js.employee_id = o.code
        LEFT JOIN LatestKRAs ks ON ks.employee_id = o.code
        LEFT JOIN uploaded_kra_kpis uk ON uk.employee_id = o.code
        GROUP BY COALESCE(NULLIF(TRIM(o.department), ''), 'Unassigned')
        ORDER BY total_employees DESC, department ASC
    """
    result = await db.execute(text(sql))
    rows = result.mappings().all()

    summary = []
    for r in rows:
        total = r["total_employees"] or 0
        completed = r["jd_completed"] or 0
        completion_rate = round((completed / total) * 100, 1) if total > 0 else 0.0
        summary.append({
            "department": r["department"],
            "total_employees": total,
            "jd_completed": completed,
            "jd_pending": total - completed,
            "pending_manager": r["pending_manager"] or 0,
            "pending_hr": r["pending_hr"] or 0,
            "in_progress": r["in_progress"] or 0,
            "not_started": r["not_started"] or 0,
            "kra_completed": r["kra_completed"] or 0,
            "completion_rate": completion_rate,
        })

    return summary


@router.get("/admin/reports/export")
async def export_admin_report(
    format: str = "excel",
    department: Optional[str] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    admin_active: str = Depends(get_current_admin),
):
    """
    Export department-wise or company-wide employee JD/KRA status report as Excel or CSV using latest sessions per employee.
    """
    from fastapi.responses import Response
    import io
    import csv

    sql = """
        WITH RankedJDs AS (
            SELECT 
                id, employee_id, status, updated_at,
                ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
            FROM jd_sessions
        ),
        LatestJDs AS (
            SELECT id, employee_id, status, updated_at
            FROM RankedJDs WHERE rn = 1
        ),
        RankedKRAs AS (
            SELECT 
                id, employee_id, status, updated_at,
                ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
            FROM kra_kpi_sessions
        ),
        LatestKRAs AS (
            SELECT id, employee_id, status, updated_at
            FROM RankedKRAs WHERE rn = 1
        )
        SELECT 
            o.code as employee_id,
            o.employee_name as name,
            e.email as email,
            o.department as department,
            o.designation as role,
            o.reporting_manager as manager_name,
            o.reporting_manager_code as manager_code,
            COALESCE(js.status, 'No JD') as jd_status,
            js.updated_at as last_active,
            CASE 
                WHEN uk.id IS NOT NULL THEN 'approved'
                ELSE COALESCE(ks.status, 'Not Started')
            END as kra_kpi_status
        FROM organogram o
        LEFT JOIN employees e ON e.id = o.code
        LEFT JOIN LatestJDs js ON js.employee_id = o.code
        LEFT JOIN LatestKRAs ks ON ks.employee_id = o.code
        LEFT JOIN uploaded_kra_kpis uk ON uk.employee_id = o.code
        WHERE 1=1
    """
    params = {}
    if department and department.strip() and department.lower() != "all":
        sql += " AND LOWER(TRIM(o.department)) = LOWER(:department)"
        params["department"] = department.strip()
    if status and status.strip() and status.lower() != "all":
        if status.lower() == "no jd":
            sql += " AND js.id IS NULL"
        else:
            sql += " AND js.status = :status"
            params["status"] = status.strip()

    sql += " ORDER BY o.department ASC, o.employee_name ASC"

    result = await db.execute(text(sql), params)
    rows = result.mappings().all()

    formatted_rows = []
    seen = set()
    for r in rows:
        emp_id = r["employee_id"]
        if emp_id in seen:
            continue
        seen.add(emp_id)

        jd_st = r["jd_status"]
        if jd_st == "approved":
            overall = "Completed"
            action_req = "None (Fully Approved)"
        elif jd_st == "sent_to_manager":
            overall = "Pending Manager Review"
            action_req = f"Pending Manager ({r['manager_name'] or r['manager_code'] or 'Manager'})"
        elif jd_st == "sent_to_hr":
            overall = "Pending HR Review"
            action_req = "Pending HR Approval"
        elif jd_st in ["collecting", "draft", "jd_generated"]:
            overall = "In Progress"
            action_req = "Employee working on draft"
        elif jd_st in ["manager_rejected", "hr_rejected"]:
            overall = "Revision Required"
            action_req = "Employee revising JD"
        else:
            overall = "Not Started"
            action_req = "Employee needs to initiate JD"

        formatted_rows.append({
            "Employee Code": emp_id,
            "Employee Name": r["name"] or "Unknown",
            "Email": r["email"] or "",
            "Department": r["department"] or "Unassigned",
            "Designation / Role": r["role"] or "Employee",
            "Reporting Manager Name": r["manager_name"] or "",
            "Reporting Manager Code": r["manager_code"] or "",
            "JD Status": jd_st,
            "KRA/KPI Status": r["kra_kpi_status"],
            "Overall Status": overall,
            "Action Required": action_req,
            "Last Active": r["last_active"].strftime("%Y-%m-%d %H:%M") if r.get("last_active") else "N/A"
        })

    dept_label = (department or "Company_Wide").replace(" ", "_").replace("/", "_")

    if format.lower() == "csv":
        output = io.StringIO()
        if formatted_rows:
            writer = csv.DictWriter(output, fieldnames=list(formatted_rows[0].keys()))
            writer.writeheader()
            writer.writerows(formatted_rows)
        else:
            output.write("Employee Code,Employee Name,Email,Department,Designation / Role,Reporting Manager Name,Reporting Manager Code,JD Status,KRA/KPI Status,Overall Status,Action Required,Last Active\n")
        
        csv_bytes = output.getvalue().encode("utf-8")
        return Response(
            content=csv_bytes,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="JD_Status_Report_{dept_label}.csv"'},
        )

    # Default to Excel export (.xlsx)
    try:
        import pandas as pd
        df_master = pd.DataFrame(formatted_rows)
        
        # Build Department Summary DataFrame
        if not df_master.empty:
            summary_records = []
            for dept_name, group_df in df_master.groupby("Department"):
                total_emp = len(group_df)
                completed = len(group_df[group_df["JD Status"] == "approved"])
                pending_mgr = len(group_df[group_df["JD Status"] == "sent_to_manager"])
                pending_hr = len(group_df[group_df["JD Status"] == "sent_to_hr"])
                in_prog = len(group_df[group_df["JD Status"].isin(["collecting", "draft", "jd_generated", "manager_rejected", "hr_rejected"])])
                not_start = len(group_df[group_df["JD Status"] == "No JD"])
                rate = round((completed / total_emp) * 100, 1) if total_emp > 0 else 0.0

                summary_records.append({
                    "Department": dept_name,
                    "Total Employees": total_emp,
                    "JD Completed": completed,
                    "JD Pending": total_emp - completed,
                    "Pending Manager": pending_mgr,
                    "Pending HR": pending_hr,
                    "In Progress / Draft": in_prog,
                    "Not Started": not_start,
                    "Completion Rate %": f"{rate}%"
                })
            df_summary = pd.DataFrame(summary_records)
        else:
            df_summary = pd.DataFrame()

        output_buf = io.BytesIO()
        with pd.ExcelWriter(output_buf, engine="openpyxl") as writer:
            if not df_summary.empty:
                df_summary.to_excel(writer, sheet_name="Department Summary", index=False)
            df_master.to_excel(writer, sheet_name="Employee Status Report", index=False)

            wb = writer.book
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            thin_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9")
            )

            fill_working = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            font_working = Font(name="Segoe UI", size=10, bold=True, color="276A3C")

            fill_review = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            font_review = Font(name="Segoe UI", size=10, bold=True, color="C65911")

            fill_progress = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            font_progress = Font(name="Segoe UI", size=10, bold=True, color="8A6D3B")

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.views.sheetView[0].showGridLines = True

                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border
                ws.row_dimensions[1].height = 28

                for row_idx in range(2, ws.max_row + 1):
                    ws.row_dimensions[row_idx].height = 22
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        cell.font = Font(name="Segoe UI", size=10)
                        cell.alignment = Alignment(vertical="center")

                        val_str = str(cell.value or "")
                        if val_str in ["approved", "Completed", "Approved", "Yes"]:
                            cell.fill = fill_working
                            cell.font = font_working
                        elif "Pending" in val_str or "sent_to_" in val_str or "Review" in val_str:
                            cell.fill = fill_review
                            cell.font = font_review
                        elif "Progress" in val_str or "collecting" in val_str or "draft" in val_str:
                            cell.fill = fill_progress
                            cell.font = font_progress

                for col in ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        return Response(
            content=output_buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="JD_Status_Report_{dept_label}.xlsx"'},
        )
    except Exception as e:
        logger.error(f"[EXPORT REPORT] Excel export failed: {e}")
        output = io.StringIO()
        if formatted_rows:
            writer = csv.DictWriter(output, fieldnames=list(formatted_rows[0].keys()))
            writer.writeheader()
            writer.writerows(formatted_rows)
        csv_bytes = output.getvalue().encode("utf-8")
        return Response(
            content=csv_bytes,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="JD_Status_Report_{dept_label}.csv"'},
        )


@router.post("/admin/kra-kpi/upload")
async def upload_kra_kpi_document(
    file: UploadFile = File(...),
    employee_id: str = Form(...),
    employee_name: str = Form(...),
    db: AsyncSession = Depends(get_db),
    admin_role: str = Depends(get_current_admin),
):
    """
    Upload and parse existing KRA/KPI document for an employee.
    Supports DOCX, PDF, and Excel (.xlsx, .xls) files.
    Auto-creates JDSession if missing, and creates a confirmed KRAKPISession.
    """

    # Check if target employee has an approved JD session
    jd_res = await db.execute(
        select(JDSession).where(
            JDSession.employee_id == employee_id,
            JDSession.status == "approved"
        )
    )
    jd_session = jd_res.scalars().first()
    if not jd_session:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "MISSING_JD",
                "message": f"Employee {employee_name} ({employee_id}) does not have an approved Job Description yet. Please prepare/approve the JD first."
            }
        )

    allowed_types = {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
        "application/msword": "docx",  # Coerce .doc to .docx parser if needed, or fallback
        "application/pdf": "pdf",
        # Excel formats
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
        "application/vnd.ms-excel": "xls",
        "application/octet-stream": None,  # Some browsers send xlsx as this
    }

    content_type = file.content_type or ""
    
    # Detect Excel by filename extension if content_type is ambiguous
    fname = (file.filename or "").lower()
    if content_type == "application/octet-stream":
        if fname.endswith(".xlsx"):
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif fname.endswith(".xls"):
            content_type = "application/vnd.ms-excel"
        else:
            raise HTTPException(
                status_code=400,
                detail="Could not determine file type. Please upload DOCX, PDF, or Excel (.xlsx/.xls).",
            )

    if content_type not in allowed_types or allowed_types[content_type] is None:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Accepted: DOCX, PDF, Excel (.xlsx/.xls). Got: {content_type}",
        )

    file_type = allowed_types[content_type]
    file_content = await file.read()

    # Validate size (10MB max)
    if len(file_content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large. Maximum 10MB")

    try:
        result = await process_kra_kpi_document(
            db=db,
            file_bytes=file_content,
            filename=file.filename,
            file_type=file_type,
            employee_id=employee_id,
            employee_name=employee_name,
            admin_role=admin_role,
        )
        return {
            "status": "success",
            "message": "KRA/KPI framework uploaded, parsed, and confirmed successfully",
            "data": result,
        }
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"[UPLOAD KRA] Processing failed: {str(e)}")
        raise HTTPException(
            status_code=400,
            detail=f"Failed to process KRA/KPI document: {str(e)}",
        )


class AnalyzePasteRequest(BaseModel):
    employee_id: str
    employee_name: str
    content: str


class ConfirmPasteRequest(BaseModel):
    employee_id: str
    employee_name: str
    jd: dict
    kra_kpi: dict


@router.post("/admin/kra-kpi/analyze-paste")
async def analyze_kra_kpi_paste_endpoint(
    request: AnalyzePasteRequest,
    db: AsyncSession = Depends(get_db),
    admin_role: str = Depends(get_current_admin),
):
    """
    Directly analyze pasted KRA/KPI raw text and return structured preview before confirmation.
    """
    # Check if target employee has an approved JD session
    jd_res = await db.execute(
        select(JDSession).where(
            JDSession.employee_id == request.employee_id,
            JDSession.status == "approved"
        )
    )
    jd_session = jd_res.scalars().first()
    if not jd_session:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "MISSING_JD",
                "message": f"Employee {request.employee_name} ({request.employee_id}) does not have an approved Job Description yet. Please prepare/approve the JD first."
            }
        )

    from app.services.kra_kpi_service import analyze_kra_kpi_text
    try:
        result = await analyze_kra_kpi_text(
            employee_id=request.employee_id,
            employee_name=request.employee_name,
            content=request.content,
        )
        return {
            "status": "success",
            "data": result,
        }
    except Exception as e:
        logger.error(f"[PASTE KRA] Analysis failed: {str(e)}")
        raise HTTPException(
            status_code=400,
            detail=f"Failed to analyze pasted KRA/KPI content: {str(e)}",
        )


@router.post("/admin/kra-kpi/confirm-paste")
async def confirm_kra_kpi_paste_endpoint(
    request: ConfirmPasteRequest,
    db: AsyncSession = Depends(get_db),
    admin_role: str = Depends(get_current_admin),
):
    """
    Save the confirmed parsed KRA/KPI and inferred JD to the employee's active session.
    """
    # Check if target employee has an approved JD session
    jd_res = await db.execute(
        select(JDSession).where(
            JDSession.employee_id == request.employee_id,
            JDSession.status == "approved"
        )
    )
    jd_session = jd_res.scalars().first()
    if not jd_session:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "MISSING_JD",
                "message": f"Employee {request.employee_name} ({request.employee_id}) does not have an approved Job Description yet. Please prepare/approve the JD first."
            }
        )

    from app.services.kra_kpi_service import save_kra_kpi_from_paste
    try:
        result = await save_kra_kpi_from_paste(
            db=db,
            employee_id=request.employee_id,
            employee_name=request.employee_name,
            jd_data=request.jd,
            kra_kpi_data=request.kra_kpi,
            admin_role=admin_role,
        )
        return {
            "status": "success",
            "message": "KRA/KPI framework confirmed and saved successfully to employee dashboard",
            "data": result,
        }
    except Exception as e:
        logger.error(f"[PASTE KRA] Save failed: {str(e)}")
        raise HTTPException(
            status_code=400,
            detail=f"Failed to confirm KRA/KPI paste: {str(e)}",
        )


class UpdateUploadedKRARequest(BaseModel):
    kras: dict


@router.put("/admin/kra-kpi/{employee_id}")
async def update_admin_kra_kpi(
    employee_id: str,
    request: UpdateUploadedKRARequest,
    db: AsyncSession = Depends(get_db),
    admin_role: str = Depends(get_current_admin),
):
    """
    Update KRA/KPI framework for an employee on the admin side.
    Updates UploadedKRAKPI if exists, otherwise updates KRAKPISession.
    """
    from app.models.kra_kpi_model import UploadedKRAKPI, KRAKPISession
    from app.core.cache import invalidate_pattern
    import logging
    
    logger = logging.getLogger(__name__)
    updated_any = False
    
    # 1. Try to find and update UploadedKRAKPI
    uploaded_res = await db.execute(
        select(UploadedKRAKPI).where(UploadedKRAKPI.employee_id == employee_id)
    )
    uploaded = uploaded_res.scalars().first()
    if uploaded:
        uploaded.kras = request.kras
        logger.info(f"Updated UploadedKRAKPI for employee {employee_id}")
        updated_any = True

    # 2. Try to find and update KRAKPISession
    session_res = await db.execute(
        select(KRAKPISession)
        .where(KRAKPISession.employee_id == employee_id)
        .order_by(KRAKPISession.updated_at.desc())
    )
    session_record = session_res.scalars().first()
    if session_record:
        session_record.kras = request.kras
        logger.info(f"Updated KRAKPISession for employee {employee_id}")
        updated_any = True
        
    if not updated_any:
        raise HTTPException(
            status_code=404,
            detail=f"No KRA/KPI framework found for employee {employee_id}."
        )
        
    await db.commit()
    
    # Invalidate cache patterns
    await invalidate_pattern(f"jds:employee:{employee_id}")
    if session_record:
        await invalidate_pattern(f"cache:jd_detail:*{session_record.jd_session_id}*")
        
    return {
        "status": "success",
        "message": "KRA/KPI framework updated successfully",
        "data": {
            "employee_id": employee_id,
            "kras": request.kras
        }
    }


