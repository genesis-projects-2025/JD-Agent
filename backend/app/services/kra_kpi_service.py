# backend/app/services/kra_kpi_service.py
"""
KRA/KPI Service — Orchestrates the 3-step KRA/KPI selection flow:

  Step 1 (kra_selection):    Generate 6–7 KRA suggestions, employee picks 3–5
  Step 2 (kpi_selection):    For each selected KRA, generate 6–7 KPI suggestions, employee picks 3–5 per KRA
  Step 3 (weight_adjustment): Employee adjusts weights via drag-and-drop, then confirms

Prerequisites (ALL THREE must be present before Step 1):
  1. Employee JD Session  — generated JD with insights
  2. Manager JD Session   — manager's generated JD
  3. Manager KRA/KPI      — manager's confirmed or draft KRAKPISession
"""

from __future__ import annotations

import asyncio
import logging
import uuid
from datetime import datetime, timezone
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

from app.models.jd_session_model import JDSession
from app.models.kra_kpi_model import KRAKPISession
from app.models.user_model import Employee
from app.agents.kra_kpi_agent import (
    generate_kra_suggestions,
    generate_kpi_suggestions_for_kra,
)

logger = logging.getLogger(__name__)


# ── Custom Exceptions ─────────────────────────────────────────────────────────

class MissingPrerequisiteError(Exception):
    def __init__(self, missing: list[str], message: str):
        self.missing = missing
        self.message = message
        super().__init__(message)


class StepError(Exception):
    """Raised when an action is called on the wrong step."""
    pass


# ── Data Extraction Helpers ───────────────────────────────────────────────────

async def _get_manager_employee_id(db: AsyncSession, employee_id: str) -> str | None:
    """
    Retrieves the manager's employee code for a given employee from the organogram records.
    """
    result = await db.execute(select(Employee).where(Employee.id == employee_id))
    emp = result.scalar_one_or_none()
    if emp and emp.reporting_manager_code:
        return emp.reporting_manager_code
    return None


def _extract_code_from_reports_to(reports_to: str) -> str | None:
    """
    Helper function to parse and extract the manager's employee code from a reports_to
    string format like 'John Doe (MGR123)'.
    """
    import re
    match = re.search(r"\(([A-Z0-9]+)\)", reports_to)
    return match.group(1) if match else None


def _extract_jd_structured(session: JDSession) -> dict:
    """
    Retrieves the structured data representation from a JD Session record.
    Falls back to parsing the plain-text field if the structured field is empty.
    """
    structured = session.jd_structured or {}
    if not structured and session.jd_text:
        import json
        try:
            parsed = json.loads(session.jd_text)
            structured = parsed.get("jd_structured_data", parsed)
        except Exception:
            pass
    return structured


def _extract_employee_data(session: JDSession) -> dict:
    """
    Extracts relevant employee JD context attributes (tasks, workflows, skills, tools)
    to form a formatted context dictionary suitable for feeding into KRA/KPI Gemini agent prompts.
    """
    insights = session.insights or {}
    structured = _extract_jd_structured(session)
    identity = insights.get("identity_context", {})

    title = (
        identity.get("title")
        or structured.get("employee_information", {}).get("title")
        or session.title or ""
    )
    department = (
        identity.get("department")
        or structured.get("employee_information", {}).get("department")
        or session.department or ""
    )
    return {
        "title": title,
        "department": department,
        "purpose": insights.get("purpose") or structured.get("purpose") or "",
        "responsibilities": structured.get("responsibilities") or [],
        "priority_tasks": insights.get("priority_tasks") or [],
        "workflows": insights.get("workflows") or {},
        "skills": insights.get("skills") or structured.get("skills") or [],
        "tools": insights.get("tools") or structured.get("tools") or [],
    }


def _extract_manager_jd_data(session: JDSession) -> dict:
    """
    Extracts high-level responsibilities and title from the manager's JD
    to align the employee's KRA suggestions with the manager's targets.
    """
    structured = _extract_jd_structured(session)
    insights = session.insights or {}
    identity = insights.get("identity_context", {})
    return {
        "title": identity.get("title") or structured.get("employee_information", {}).get("title") or session.title or "",
        "department": identity.get("department") or session.department or "",
        "responsibilities": structured.get("responsibilities") or [],
    }


# ── Prerequisite Check ────────────────────────────────────────────────────────

COMPLETED_JD_STATUSES = {
    "jd_generated", "sent_to_manager", "manager_approved", "sent_to_hr",
    "hr_approved", "approved", "completed", "manager_rejected", "hr_rejected",
}


async def check_prerequisites(
    db: AsyncSession, jd_session_id: str, employee_id: str, bypass_manager: bool = False,
) -> dict[str, Any]:
    """
    Validates that all prerequisites for KRA/KPI generation are satisfied:
    1. Employee's JD must exist and be approved by the manager.
    2. Manager's JD must be generated and approved.
    3. Manager's KRA/KPI framework must be created and confirmed/drafted.

    Executive Bypass Rule:
    High-level roles (Director, VP, CEO, MD, President) reporting directly to managing structures
    do not require manager JDs/KRAs. If the employee holds an executive title or level,
    the manager validation checks are bypassed automatically.
    """
    missing = []
    details = {}

    # ── Executive Bypass Check (Employee Level) ──
    # Check if the employee holding the session is high-level/executive
    try:
        from sqlalchemy import text
        res = await db.execute(
            text("SELECT designation, joblevel FROM organogram WHERE code = :code"),
            {"code": employee_id}
        )
        emp_org = res.mappings().first()
        if emp_org:
            emp_desig = (emp_org.get("designation") or "").lower()
            emp_level = str(emp_org.get("joblevel") or "").lower()
            if any(k in emp_desig for k in ["director", "president", "vp", "vice president", "ceo", "coo", "md", "board", "md & ceo"]) or "exec" in emp_level or "l1" in emp_level or "l2" in emp_level:
                bypass_manager = True
                logger.info(f"[ExecutiveBypass] Employee {employee_id} ({emp_desig}, level {emp_level}) bypassed manager prerequisites.")
    except Exception as e:
        logger.error(f"Error checking employee designation in organogram: {e}")

    # 1. Employee JD
    emp_result = await db.execute(
        select(JDSession).where(JDSession.id == uuid.UUID(jd_session_id))
    )
    employee_session = emp_result.scalar_one_or_none()

    if not employee_session:
        missing.append("employee_jd")
        details["employee_jd"] = "Employee JD session not found."
    elif not employee_session.jd_text and not employee_session.jd_structured:
        missing.append("employee_jd")
        details["employee_jd"] = (
            "Employee JD has not been generated yet. "
            "Complete the interview and generate the JD first."
        )
    else:
        # Option 1: Backwards compatibility - check if they already have KRA/KPI data
        kra_res = await db.execute(
            select(KRAKPISession).where(KRAKPISession.jd_session_id == str(jd_session_id))
        )
        kra_session = kra_res.scalar_one_or_none()
        is_legacy = kra_session is not None and kra_session.kras and len(kra_session.kras) > 0

        if not bypass_manager and not is_legacy:
            # Enforce that the JD is approved by the manager (status is sent_to_hr or approved)
            if employee_session.status not in {"sent_to_hr", "approved"}:
                missing.append("employee_jd_approved")
                details["employee_jd_approved"] = (
                    "Your Job Description has not been approved by your manager yet. "
                    "Your manager must review and approve your JD before you can generate your KRA/KPI performance goals."
                )

    # 2. Manager ID
    manager_employee_id = None
    manager_jd_session = None
    manager_kra_session = None

    if not bypass_manager:
        manager_employee_id = await _get_manager_employee_id(db, employee_id)
        if not manager_employee_id and employee_session and employee_session.insights:
            ic = employee_session.insights.get("identity_context", {})
            manager_employee_id = _extract_code_from_reports_to(ic.get("reports_to", ""))

        if manager_employee_id:
            # Audit manager's designation in organogram to auto-bypass executive manager requirements
            from sqlalchemy import text
            try:
                res = await db.execute(
                    text("SELECT designation, code FROM organogram WHERE code = :code"),
                    {"code": manager_employee_id}
                )
                mgr_org = res.mappings().first()
                if mgr_org:
                    mgr_desig = (mgr_org.get("designation") or "").lower()
                    mgr_code = (mgr_org.get("code") or "").upper()
                    if any(k in mgr_desig for k in ["director", "president", "vp", "vice president", "ceo", "coo", "md", "board", "md & ceo"]) or mgr_code.startswith("DIR") or mgr_code.startswith("MD"):
                        bypass_manager = True
            except Exception as e:
                logger.error(f"Error checking manager designation in organogram: {e}")

        if not bypass_manager:
            if not manager_employee_id:
                missing += ["manager_jd", "manager_kra_kpi"]
                details["manager_jd"] = "Could not identify your reporting manager. Please contact HR to update your reporting structure."
                details["manager_kra_kpi"] = "Manager not identified — cannot check manager KRA/KPI."
                raise MissingPrerequisiteError(missing, _build_missing_message(details))

            # 3. Manager JD
            mgr_jd_result = await db.execute(
                select(JDSession)
                .where(JDSession.employee_id == manager_employee_id)
                .where(JDSession.status.in_(list(COMPLETED_JD_STATUSES)))
                .order_by(JDSession.updated_at.desc())
            )
            manager_jd_session = mgr_jd_result.scalars().first()

            if not manager_jd_session:
                missing.append("manager_jd")
                details["manager_jd"] = (
                    f"Your manager's Job Description has not been generated yet. "
                    f"Request your manager (ID: {manager_employee_id}) to complete their JD interview first."
                )

            # 4. Manager KRA/KPI
            mgr_kra_result = await db.execute(
                select(KRAKPISession)
                .where(KRAKPISession.employee_id == manager_employee_id)
                .where(KRAKPISession.status.in_(["confirmed", "draft"]))
                .order_by(KRAKPISession.updated_at.desc())
            )
            manager_kra_session = mgr_kra_result.scalars().first()

            if not manager_kra_session:
                from app.models.kra_kpi_model import UploadedKRAKPI
                uploaded_res = await db.execute(
                    select(UploadedKRAKPI).where(UploadedKRAKPI.employee_id == manager_employee_id)
                )
                manager_uploaded_kra = uploaded_res.scalars().first()
                if manager_uploaded_kra:
                    class MockKRAKPISession:
                        def __init__(self, uploaded):
                            self.id = uploaded.id
                            self.kras = uploaded.kras
                    manager_kra_session = MockKRAKPISession(manager_uploaded_kra)

            if not manager_kra_session:
                missing.append("manager_kra_kpi")
                details["manager_kra_kpi"] = (
                    f"Your manager's KRA/KPI framework has not been created yet. "
                    f"Request your manager (ID: {manager_employee_id}) to generate their KRAs/KPIs first."
                )
    else:
        # Bypassed manager prerequisites, but let's try to fetch them optionally as reference anyway if they happen to exist!
        try:
            manager_employee_id = await _get_manager_employee_id(db, employee_id)
            if not manager_employee_id and employee_session and employee_session.insights:
                ic = employee_session.insights.get("identity_context", {})
                manager_employee_id = _extract_code_from_reports_to(ic.get("reports_to", ""))
            
            if manager_employee_id:
                mgr_jd_result = await db.execute(
                    select(JDSession)
                    .where(JDSession.employee_id == manager_employee_id)
                    .where(JDSession.status.in_(list(COMPLETED_JD_STATUSES)))
                    .order_by(JDSession.updated_at.desc())
                )
                manager_jd_session = mgr_jd_result.scalars().first()

                mgr_kra_result = await db.execute(
                    select(KRAKPISession)
                    .where(KRAKPISession.employee_id == manager_employee_id)
                    .where(KRAKPISession.status.in_(["confirmed", "draft"]))
                    .order_by(KRAKPISession.updated_at.desc())
                )
                manager_kra_session = mgr_kra_result.scalars().first()

                if not manager_kra_session:
                    from app.models.kra_kpi_model import UploadedKRAKPI
                    uploaded_res = await db.execute(
                        select(UploadedKRAKPI).where(UploadedKRAKPI.employee_id == manager_employee_id)
                    )
                    manager_uploaded_kra = uploaded_res.scalars().first()
                    if manager_uploaded_kra:
                        class MockKRAKPISession:
                            def __init__(self, uploaded):
                                self.id = uploaded.id
                                self.kras = uploaded.kras
                        manager_kra_session = MockKRAKPISession(manager_uploaded_kra)
        except Exception:
            pass

    if missing:
        raise MissingPrerequisiteError(missing, _build_missing_message(details))

    return {
        "employee_session": employee_session,
        "manager_employee_id": manager_employee_id,
        "manager_jd_session": manager_jd_session,
        "manager_kra_session": manager_kra_session,
    }


def _build_missing_message(details: dict) -> str:
    lines = ["❌ KRA/KPI generation requires the following missing information:\n"]
    if "employee_jd" in details:
        lines.append(f"📄 Employee JD: {details['employee_jd']}")
    if "manager_jd" in details:
        lines.append(f"👔 Manager JD: {details['manager_jd']}")
    if "manager_kra_kpi" in details:
        lines.append(f"🎯 Manager KRA/KPI: {details['manager_kra_kpi']}")
    lines.append("\nOnce all three are available, KRA/KPI generation will proceed automatically.")
    return "\n".join(lines)


# ── DB Helpers ────────────────────────────────────────────────────────────────

async def get_kra_kpi_by_jd_session(
    db: AsyncSession, jd_session_id: str, employee_id: str | None = None,
) -> KRAKPISession | None:
    from sqlalchemy.orm import selectinload

    if employee_id:
        # 1. Primary: Exact match on BOTH employee_id and jd_session_id
        result = await db.execute(
            select(KRAKPISession)
            .where(
                KRAKPISession.employee_id == employee_id,
                KRAKPISession.jd_session_id == jd_session_id
            )
            .options(selectinload(KRAKPISession.conversation_turns))
            .order_by(KRAKPISession.updated_at.desc())
        )
        rec = result.scalars().first()
        if rec:
            return rec

        # 2. Secondary: Match by employee_id alone (handles cloned / role-template JDs)
        result_emp = await db.execute(
            select(KRAKPISession)
            .where(KRAKPISession.employee_id == employee_id)
            .options(selectinload(KRAKPISession.conversation_turns))
            .order_by(KRAKPISession.updated_at.desc())
        )
        rec_emp = result_emp.scalars().first()
        if rec_emp:
            return rec_emp

        # If employee_id was explicitly supplied but no session exists for THIS employee,
        # return None to prevent leaking another employee's session!
        return None

    # Fallback if employee_id was not specified
    result = await db.execute(
        select(KRAKPISession)
        .where(KRAKPISession.jd_session_id == jd_session_id)
        .options(selectinload(KRAKPISession.conversation_turns))
        .order_by(KRAKPISession.updated_at.desc())
    )
    return result.scalars().first()


# Cascade goal alignment filtering helper
import math
from app.services.vector_service import get_embeddings_for_text

async def filter_manager_kras_semantically(employee_tasks: list[str], manager_kras: list[dict]) -> list[dict]:
    """Filters the manager's KRAs semantically against the employee's tasks,
    returning only the KRAs that are relevant to the employee's role.
    """
    if not manager_kras or not employee_tasks:
        return manager_kras
        
    try:
        # Get employee tasks embedding (concatenate first 5 tasks to represent the role)
        emp_text = " ".join(employee_tasks[:5])
        if not emp_text.strip():
            return manager_kras
            
        emp_vector = await get_embeddings_for_text(emp_text)
        
        # Embed each manager KRA and calculate similarity
        scored_kras = []
        for kra in manager_kras:
            title = kra.get("title", "")
            desc = kra.get("description", "")
            kra_text = f"{title} {desc}".strip()
            if not kra_text:
                continue
                
            kra_vector = await get_embeddings_for_text(kra_text)
            
            # Cosine similarity
            dot_product = sum(a * b for a, b in zip(emp_vector, kra_vector))
            magnitude_a = math.sqrt(sum(a * a for a in emp_vector))
            magnitude_b = math.sqrt(sum(b * b for b in kra_vector))
            if magnitude_a == 0 or magnitude_b == 0:
                similarity = 0.0
            else:
                similarity = dot_product / (magnitude_a * magnitude_b)
                
            scored_kras.append((similarity, kra))
            
        # Sort by similarity desc
        scored_kras.sort(key=lambda x: x[0], reverse=True)
        
        # Keep top 3 or those with similarity >= 0.35
        # Ensure we return at least 1 KRA if any exist
        filtered = []
        for sim, kra in scored_kras:
            if sim >= 0.35 or len(filtered) < 2:
                filtered.append(kra)
                
        logger.info(f"[CascadeAlignment] Filtered manager KRAs from {len(manager_kras)} down to {len(filtered)} based on semantic relevance.")
        return filtered
    except Exception as e:
        logger.warning(f"Error in filter_manager_kras_semantically: {e}")
        return manager_kras


# ── Step 1: Generate KRA Suggestions ─────────────────────────────────────────

async def generate_kra_suggestions_for_employee(
    db: AsyncSession, jd_session_id: str, employee_id: str, bypass_manager: bool = False,
) -> KRAKPISession:
    """
    Step 1: Orchestrates the KRA suggestion generation phase.
    
    1. Runs prerequisite validations (checks for approved JD and manager JDs/KRAs, applying executive bypasses).
    2. Extracts employee responsibilities and applies semantic filtering to align them with manager's KRAs.
    3. Calls the Gemini-based agent (`generate_kra_suggestions`) to suggest 6-7 custom KRAs.
    4. Upserts this data into a `KRAKPISession` record in the database.
    5. Sets `generation_step` to `"kra_selection"` and status to `"draft"`.
    """
    context = await check_prerequisites(db, jd_session_id, employee_id, bypass_manager=bypass_manager)

    employee_session: JDSession = context["employee_session"]
    manager_jd_session: JDSession | None = context.get("manager_jd_session")
    manager_kra_session: KRAKPISession | None = context.get("manager_kra_session")
    manager_employee_id: str | None = context.get("manager_employee_id")

    employee_data = _extract_employee_data(employee_session)
    manager_jd_data = _extract_manager_jd_data(manager_jd_session) if manager_jd_session else {}
    manager_kras = (manager_kra_session.kras or {}).get("kras", []) if manager_kra_session else []

    # Cascade goal alignment filtering
    if manager_kras and employee_data.get("responsibilities"):
        emp_tasks = []
        for r in employee_data.get("responsibilities", []):
            if isinstance(r, dict):
                emp_tasks.append(r.get("description") or r.get("task") or "")
            else:
                emp_tasks.append(str(r))
        manager_kras = await filter_manager_kras_semantically(emp_tasks, manager_kras)

    logger.info(f"[KRAKPIService] Step 1: Generating KRA suggestions for employee={employee_id}")

    kra_payload = await generate_kra_suggestions(
        employee_data=employee_data,
        manager_jd_data=manager_jd_data,
        manager_kras_data=manager_kras,
    )

    now = datetime.now(timezone.utc)

    # Upsert
    existing = await get_kra_kpi_by_jd_session(db, jd_session_id)
    if existing:
        existing.kra_suggestions = kra_payload
        existing.selected_kra_ids = None
        existing.kpi_suggestions = None
        existing.selected_kpi_ids = None
        existing.kras = None
        existing.generation_step = "kra_selection"
        existing.status = "draft"
        existing.manager_employee_id = manager_employee_id
        existing.manager_jd_session_id = str(manager_jd_session.id) if manager_jd_session else None
        existing.manager_kra_kpi_session_id = str(manager_kra_session.id) if manager_kra_session else None
        existing.generation_model = "gemini-2.5-flash"
        existing.generation_error = None
        existing.generated_at = now
        existing.updated_at = now
        record = existing
    else:
        record = KRAKPISession(
            id=uuid.uuid4(),
            jd_session_id=jd_session_id,
            employee_id=employee_id,
            manager_employee_id=manager_employee_id,
            manager_jd_session_id=str(manager_jd_session.id) if manager_jd_session else None,
            manager_kra_kpi_session_id=str(manager_kra_session.id) if manager_kra_session else None,
            kra_suggestions=kra_payload,
            generation_step="kra_selection",
            status="draft",
            generation_model="gemini-2.5-flash",
            generated_at=now,
        )
        db.add(record)

    await db.commit()
    await db.refresh(record)
    return record


# ── Step 2: Select KRAs → Generate KPI Suggestions ───────────────────────────

async def select_kras_and_generate_kpis(
    db: AsyncSession,
    jd_session_id: str,
    selected_kra_ids: list[str],
    employee_id: str | None = None,
) -> KRAKPISession:
    """
    Step 2: Handles employee selection of KRAs and generates KPI suggestions.
    """

    # Validation: at least 1 KRA must be selected
    if len(selected_kra_ids) < 1:
        raise StepError("Please select at least 1 KRA.")

    record = await get_kra_kpi_by_jd_session(db, jd_session_id, employee_id=employee_id)
    if not record:
        raise StepError("No KRA/KPI session found. Generate KRA suggestions first.")
    if record.generation_step not in ("kra_selection", "kpi_selection"):
        raise StepError(f"Cannot select KRAs in step: {record.generation_step}")

    # Resolve the full KRA objects for selected IDs
    all_suggestions = (record.kra_suggestions or {}).get("kra_suggestions", [])
    kra_map = {k["kra_id"]: k for k in all_suggestions}
    selected_kras = [kra_map[kid] for kid in selected_kra_ids if kid in kra_map]

    if not selected_kras:
        raise StepError("None of the selected KRA IDs match generated suggestions.")

    # Load employee data for KPI generation
    emp_session_result = await db.execute(
        select(JDSession).where(JDSession.id == uuid.UUID(record.jd_session_id))
    )
    emp_session = emp_session_result.scalar_one_or_none()
    if not emp_session:
        raise StepError("Employee JD session not found.")
    employee_data = _extract_employee_data(emp_session)

    # Fetch most recent skill ratings of this employee to identify skill gaps
    skill_gaps = []
    try:
        from app.models.kra_kpi_model import KRAKPISession as _KRAKPISession
        ratings_res = await db.execute(
            select(_KRAKPISession.skill_ratings)
            .where(
                _KRAKPISession.employee_id == record.employee_id,
                _KRAKPISession.skill_ratings.isnot(None),
            )
            .order_by(_KRAKPISession.updated_at.desc())
            .limit(1)
        )
        ratings_val = ratings_res.scalar_one_or_none()
        if ratings_val:
            if isinstance(ratings_val, list):
                for item in ratings_val:
                    if isinstance(item, dict):
                        name = item.get("name") or item.get("skill")
                        rating = item.get("rating")
                        if name and rating is not None:
                            try:
                                if float(rating) < 6.0:
                                    skill_gaps.append(f"{name} (rating: {rating}/10)")
                            except ValueError:
                                pass
            elif isinstance(ratings_val, dict):
                for name, rating in ratings_val.items():
                    if rating is not None:
                        try:
                            if float(rating) < 6.0:
                                skill_gaps.append(f"{name} (rating: {rating}/10)")
                        except ValueError:
                            pass
    except Exception as e:
        logger.warning(f"Failed to fetch employee skill gaps for KPI generation: {e}")

    logger.info(
        f"[KRAKPIService] Step 2: Generating KPI suggestions for {len(selected_kras)} KRAs in parallel"
    )

    # Generate KPI suggestions for all selected KRAs in parallel
    kpi_tasks = [
        generate_kpi_suggestions_for_kra(kra=kra, employee_data=employee_data, skill_gaps=skill_gaps)
        for kra in selected_kras
    ]
    results = await asyncio.gather(*kpi_tasks)

    # Index by kra_id
    kpi_suggestions = {r["kra_id"]: r for r in results}

    now = datetime.now(timezone.utc)
    record.selected_kra_ids = selected_kra_ids
    record.kpi_suggestions = kpi_suggestions
    record.selected_kpi_ids = None
    record.kras = None
    record.generation_step = "kpi_selection"
    record.updated_at = now

    await db.commit()
    await db.refresh(record)
    return record


# ── Step 3: Select KPIs → Build Final Weight-Adjustment Payload ───────────────

async def select_kpis_and_build_final(
    db: AsyncSession,
    jd_session_id: str,
    selected_kpi_ids: dict[str, list[str]],
    employee_id: str | None = None,
) -> KRAKPISession:
    """
    Step 3a: Handles employee selection of KPIs for each chosen KRA.
    """
    record = await get_kra_kpi_by_jd_session(db, jd_session_id, employee_id=employee_id)
    if not record:
        raise StepError("No KRA/KPI session found.")
    if record.generation_step not in ("kpi_selection", "weight_adjustment"):
        raise StepError(f"Cannot select KPIs in step: {record.generation_step}")

    # Validate KPI counts per KRA — at least 1 KPI must be selected per KRA
    for kra_id, kpi_ids in selected_kpi_ids.items():
        if len(kpi_ids) < 1:
            raise StepError(
                f"Select at least 1 KPI for each KRA. "
                f"KRA '{kra_id}' has no KPIs selected."
            )

    # Resolve full KRA objects
    all_suggestions = (record.kra_suggestions or {}).get("kra_suggestions", [])
    kra_map = {k["kra_id"]: k for k in all_suggestions}

    kpi_suggestion_map = record.kpi_suggestions or {}

    selected_kra_ids = record.selected_kra_ids or []

    # Build final KRAs list — weights start as None, employee sets them in Step 3
    final_kras = []
    for i, kra_id in enumerate(selected_kra_ids):
        kra_base = kra_map.get(kra_id, {})
        kpi_ids = selected_kpi_ids.get(kra_id, [])

        # Get selected KPI full objects
        kpi_bank = kpi_suggestion_map.get(kra_id, {}).get("kpi_suggestions", [])
        kpi_obj_map = {k["kpi_id"]: k for k in kpi_bank}
        selected_kpis = [kpi_obj_map[kid] for kid in kpi_ids if kid in kpi_obj_map]

        final_kras.append({
            "kra_id": kra_id,
            "title": kra_base.get("title", ""),
            "description": kra_base.get("description", ""),
            "source_tasks": kra_base.get("source_tasks", []),
            "weight": None,  # Employee assigns weights manually in Step 3
            "manager_impact": kra_base.get("manager_impact", ""),
            "kpis": selected_kpis,
        })

    now = datetime.now(timezone.utc)
    record.selected_kpi_ids = selected_kpi_ids
    record.kras = {"kras": final_kras, "total_weight": 0}
    record.generation_step = "weight_adjustment"
    record.updated_at = now

    await db.commit()
    await db.refresh(record)
    return record


# ── Step 4: Save Weight Adjustments and Confirm ───────────────────────────────

async def save_weights_and_confirm(
    db: AsyncSession,
    jd_session_id: str,
    kras_with_weights: list[dict],
    confirm: bool = False,
    employee_id: str | None = None,
) -> KRAKPISession:
    """
    Step 3b: Save weight configurations set by the employee for their selected KRAs and KPIs.
    """
    total = sum(k.get("weight", 0) for k in kras_with_weights)
    if abs(total - 100) > 1:  # Allow ±1 for rounding
        raise StepError(f"KRA weights must sum to 100. Current total: {total}")

    # Validate KPI weights sum to 100 for each KRA
    for kra in kras_with_weights:
        kpis = kra.get("kpis", [])
        if kpis:
            kpi_total = sum(kp.get("weight", 0) for kp in kpis)
            if abs(kpi_total - 100) > 1:
                raise StepError(f"KPI weights for KRA '{kra.get('title')}' must sum to 100. Current total: {kpi_total}")
            
            # Normalize KPI weights to exactly 100
            if kpi_total != 100:
                diff = 100 - kpi_total
                kpis[-1]["weight"] = kpis[-1]["weight"] + diff

    record = await get_kra_kpi_by_jd_session(db, jd_session_id, employee_id=employee_id)
    if not record:
        raise StepError("No KRA/KPI session found.")

    # Normalize KRA weights to exactly 100
    if total != 100:
        diff = 100 - total
        kras_with_weights[-1]["weight"] = kras_with_weights[-1]["weight"] + diff

    now = datetime.now(timezone.utc)
    record.kras = {"kras": kras_with_weights, "total_weight": 100}
    record.updated_at = now

    if confirm:
        record.generation_step = "confirmed"
        record.status = "confirmed"
        record.confirmed_at = now

    await db.commit()
    await db.refresh(record)

    return record



def _is_structured_template(rows: list) -> tuple[bool, dict]:
    """
    Detect if the Excel file matches our structured KRA/KPI bulk upload template.
    Template format has columns: Employee_ID, Employee_Name, KRA_Title, KRA_Weight_%, KPI_Title, ...
    Returns (is_template, column_index_map).
    """
    if not rows:
        return False, {}
    header_row = None
    header_idx = -1
    for i, row in enumerate(rows[:5]):
        cells = [str(c).lower().strip() if c is not None else "" for c in row]
        # Must have employee_id and kra_title (or kra) columns
        has_emp_id = any("employee_id" in c or "emp_id" in c or "employee id" in c for c in cells)
        has_kra = any("kra_title" in c or "kra title" in c for c in cells)
        if has_emp_id and has_kra:
            header_row = cells
            header_idx = i
            break
    if header_row is None:
        return False, {}
    col_map = {}
    for idx, cell in enumerate(header_row):
        if "employee_id" in cell or "emp_id" in cell or "employee id" in cell:
            col_map["employee_id"] = idx
        elif "employee_name" in cell or "emp_name" in cell or "employee name" in cell:
            col_map["employee_name"] = idx
        elif "kra_title" in cell or "kra title" in cell:
            col_map["kra_title"] = idx
        elif "kra_weight" in cell or "kra weight" in cell or "weight" in cell:
            col_map["kra_weight"] = idx
        elif "kpi_title" in cell or "kpi title" in cell:
            col_map["kpi_title"] = idx
        elif "kpi_target" in cell or "kpi target" in cell or "target_date" in cell or "target date" in cell:
            col_map["kpi_target"] = idx
        elif "kpi_description" in cell or "kpi description" in cell or "description" in cell:
            col_map["kpi_description"] = idx
    col_map["_header_row_idx"] = header_idx
    return True, col_map


def parse_kra_kpi_excel_bulk(file_bytes: bytes, file_type_lower: str) -> dict[str, list[dict]]:
    """
    Parse structured template Excel files into a dict keyed by employee_id.
    Each value is a list of KRA dicts: [{"title": ..., "description": ..., "weight": ..., "kpis": [...]}]
    Handles multi-employee files.
    """
    import io
    rows = []

    if file_type_lower == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            # Try to find the right sheet (KRA_KPI first, else active)
            sheet_name = "KRA_KPI" if "KRA_KPI" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()
        except ImportError:
            raise Exception("openpyxl is required to parse .xlsx files.")
    elif file_type_lower == "xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            ws = wb.sheet_by_index(0)
            for r_idx in range(ws.nrows):
                rows.append(ws.row_values(r_idx))
        except ImportError:
            raise Exception("xlrd is required to parse .xls files.")
    else:
        raise Exception("Unsupported file type for bulk Excel parsing")

    is_template, col_map = _is_structured_template(rows)
    if not is_template:
        raise ValueError("File does not match the structured KRA/KPI bulk upload template format.")

    header_idx = col_map.get("_header_row_idx", 0)
    emp_id_col = col_map.get("employee_id", 0)
    emp_name_col = col_map.get("employee_name", 1)
    kra_title_col = col_map.get("kra_title", 2)
    kra_weight_col = col_map.get("kra_weight", 3)
    kpi_title_col = col_map.get("kpi_title", 4)
    kpi_target_col = col_map.get("kpi_target", 5)
    kpi_desc_col = col_map.get("kpi_description", 6)

    # employee_id -> {"name": str, "kras": {kra_title_lower: kra_dict}}
    employees: dict[str, dict] = {}

    def get_val(row, col_idx):
        if col_idx is not None and 0 <= col_idx < len(row):
            v = row[col_idx]
            if v is None:
                return ""
            return str(v).strip()
        return ""

    for row in rows[header_idx + 1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue

        emp_id = get_val(row, emp_id_col)
        emp_name = get_val(row, emp_name_col)
        kra_title = get_val(row, kra_title_col)
        kpi_title = get_val(row, kpi_title_col)

        if not emp_id or not kra_title or not kpi_title:
            continue

        # Parse weight — could be "25", "25%", or 0.25 float
        raw_weight = get_val(row, kra_weight_col)
        kra_weight = None
        if raw_weight:
            try:
                w = float(raw_weight.replace("%", "").strip())
                # If weight is stored as decimal (e.g. 0.25), convert to %
                kra_weight = round(w * 100) if w <= 1.0 else round(w)
            except ValueError:
                kra_weight = None

        kpi_target = get_val(row, kpi_target_col)
        kpi_desc = get_val(row, kpi_desc_col)

        if emp_id not in employees:
            employees[emp_id] = {"name": emp_name, "kras": {}}

        kra_key = kra_title.lower().strip()
        if kra_key not in employees[emp_id]["kras"]:
            employees[emp_id]["kras"][kra_key] = {
                "title": kra_title,
                "description": "",
                "weight": kra_weight,
                "kpis": [],
            }
        elif kra_weight is not None and employees[emp_id]["kras"][kra_key]["weight"] is None:
            employees[emp_id]["kras"][kra_key]["weight"] = kra_weight

        kpi_entry = {"title": kpi_title, "description": kpi_desc}
        if kpi_target:
            kpi_entry["target_date"] = kpi_target

        employees[emp_id]["kras"][kra_key]["kpis"].append(kpi_entry)

    # Convert to final format: {employee_id: [kra_dict, ...]}
    result = {}
    for emp_id, emp_data in employees.items():
        result[emp_id] = {
            "employee_name": emp_data["name"],
            "kras": list(emp_data["kras"].values()),
        }

    if not result:
        raise Exception("No valid KRA/KPI rows found in the template. Check that Employee_ID, KRA_Title, and KPI_Title columns are filled.")

    return result


def split_kpi_text(text: str) -> list[str]:
    """
    Splits a single KPI string that contains multiple KPIs (e.g., separated by newlines,
    bullet points, or numeric lists) into individual KPI titles/descriptions.
    """
    if not text:
        return []
    import re
    # Split by:
    # 1. Newlines
    # 2. Inline numbering like ' 2. ' or ' 2) ' (space after dot/parenthesis optional)
    # 3. Inline bullets like ' • ', ' - ', ' * ' (space after bullet optional)
    parts = re.split(r'[\r\n]+|\s+(?=\d+[\.\)]\s*)|\s+[\u2022•\-\*]\s*', text)
    
    kpis = []
    # Clean up any remaining leading bullet/numbering prefixes on each part
    prefix_pattern = re.compile(r'^([•\-\*\u2022\d]+\s*[\.\)]\s*|[•\-\*\u2022]\s*)')
    for part in parts:
        part_str = part.strip()
        if not part_str:
            continue
        cleaned = prefix_pattern.sub('', part_str).strip()
        if cleaned:
            kpis.append(cleaned)
            
    if not kpis:
        trimmed = text.strip()
        if trimmed:
            kpis = [trimmed]
            
    return kpis


def split_all_kpis_in_kras(kras: list[dict]) -> list[dict]:
    """
    Post-processes a list of KRAs to ensure that if any KPI contains a list/multiple KPIs,
    it is expanded into separate KPI dictionaries.
    """
    if not kras:
        return []
    
    updated_kras = []
    for kra in kras:
        kpis_new = []
        for kpi in kra.get("kpis", []):
            title = kpi.get("title", "")
            desc = kpi.get("description", "")
            target_date = kpi.get("target_date")
            
            titles_split = split_kpi_text(title)
            descs_split = split_kpi_text(desc) if desc else []
            
            # If we split either the title or the description into multiple items,
            # expand them into separate KPI dicts
            if len(titles_split) > 1 or len(descs_split) > 1:
                num_kpis = max(len(titles_split), len(descs_split))
                for idx in range(num_kpis):
                    t = titles_split[idx] if idx < len(titles_split) else (titles_split[-1] if titles_split else "KPI")
                    d = descs_split[idx] if idx < len(descs_split) else ""
                    kpi_entry = {"title": t, "description": d}
                    if target_date:
                        kpi_entry["target_date"] = target_date
                    kpis_new.append(kpi_entry)
            else:
                kpis_new.append(kpi)
                
        kra_copy = dict(kra)
        kra_copy["kpis"] = kpis_new
        updated_kras.append(kra_copy)
        
    return updated_kras



def parse_kra_kpi_excel(file_bytes: bytes, file_type_lower: str) -> list[dict]:
    """
    Single-employee Excel parser supporting two formats:

    FORMAT A — Structured template (Employee_ID, KRA_Title, KPI_Title, ... columns):
      Detected automatically via _is_structured_template().

    FORMAT B — Srinivas-style (the real format from actual files):
      Col A: KRA title on first row of each group; blank or weight (e.g. "0.25") on subsequent rows
      Col B: KPI title — one KPI per row (first KPI is on same row as KRA title)
      Col C: (optional) target date

      Example:
        Row 1:  KRA                        | KPI           | Date
        Row 2:  KRA Title Here             | KPI 1 text    | 2026-01-20  ← KRA title + first KPI
        Row 3:  0.25  (weight decimal)     | KPI 2 text    | 2026-03-01  ← weight row + second KPI
        Row 4:  (empty)                    | KPI 3 text    |             ← more KPIs
        Row 6:  Next KRA Title             | KPI 1 text    | 2026-03-20
        ...

    Returns list of:
      {"title": str, "description": str, "weight": int|None, "kpis": [{"title": str, "description": str, "target_date": str?}]}
    """
    import io
    rows = []

    if file_type_lower == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            # Prefer sheet named "KRA_KPI" or "KRA", else use active
            preferred = ["KRA_KPI", "KRA"]
            sheet_name = next((s for s in preferred if s in wb.sheetnames), wb.sheetnames[0])
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()
        except ImportError:
            raise Exception("openpyxl is required to parse .xlsx files.")
    elif file_type_lower == "xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            ws = wb.sheet_by_index(0)
            for r_idx in range(ws.nrows):
                rows.append(ws.row_values(r_idx))
        except ImportError:
            raise Exception("xlrd is required to parse .xls files.")
    else:
        raise Exception("Unsupported file type for direct Excel parsing")

    if not rows:
        raise Exception("Excel sheet is empty")

    # ── FORMAT A: Structured template (Employee_ID + KRA_Title columns) ───────
    is_template, col_map = _is_structured_template(rows)
    if is_template:
        logger.info("[parse_kra_kpi_excel] Detected structured template format.")
        header_idx = col_map.get("_header_row_idx", 0)
        kra_title_col = col_map.get("kra_title", 2)
        kra_weight_col = col_map.get("kra_weight", 3)
        kpi_title_col = col_map.get("kpi_title", 4)
        kpi_target_col = col_map.get("kpi_target", 5)
        kpi_desc_col = col_map.get("kpi_description", 6)

        kras: list[dict] = []
        kra_index: dict[str, dict] = {}

        def get_val(row, col_idx):
            if col_idx is not None and 0 <= col_idx < len(row):
                v = row[col_idx]
                return str(v).strip() if v is not None else ""
            return ""

        for row in rows[header_idx + 1:]:
            if not any(c is not None and str(c).strip() for c in row):
                continue
            kra_title = get_val(row, kra_title_col)
            kpi_title = get_val(row, kpi_title_col)
            if not kra_title or not kpi_title:
                continue

            raw_weight = get_val(row, kra_weight_col)
            kra_weight = None
            if raw_weight:
                try:
                    w = float(raw_weight.replace("%", "").strip())
                    kra_weight = round(w * 100) if w <= 1.0 else round(w)
                except ValueError:
                    pass

            kpi_desc = get_val(row, kpi_desc_col)
            kpi_target = get_val(row, kpi_target_col)

            key = kra_title.lower().strip()
            if key not in kra_index:
                kra_dict: dict = {"title": kra_title, "description": "", "weight": kra_weight, "kpis": []}
                kras.append(kra_dict)
                kra_index[key] = kra_dict
            elif kra_weight is not None and kra_index[key]["weight"] is None:
                kra_index[key]["weight"] = kra_weight

            # Split kpi_title and kpi_desc and loop max times to capture all split KPIs
            kpi_titles = split_kpi_text(kpi_title)
            kpi_descs = split_kpi_text(kpi_desc) if kpi_desc else []
            
            num_kpis = max(len(kpi_titles), len(kpi_descs))
            for idx_kpi in range(num_kpis):
                if idx_kpi < len(kpi_titles):
                    kt = kpi_titles[idx_kpi]
                elif kpi_titles:
                    kt = kpi_titles[-1]
                else:
                    kt = "KPI"
                    
                if idx_kpi < len(kpi_descs):
                    kd = kpi_descs[idx_kpi]
                else:
                    kd = ""
                    
                kpi_entry: dict = {"title": kt, "description": kd}
                if kpi_target:
                    kpi_entry["target_date"] = kpi_target
                kra_index[key]["kpis"].append(kpi_entry)

        if kras:
            return kras
        # Fall through to Format B

    # ── FORMAT B: Srinivas-style (Col A = KRA or weight; Col B = KPI; Col C = date) ──
    logger.info("[parse_kra_kpi_excel] Using Srinivas-style format parser.")

    def _get(row, idx):
        """Safe cell getter — strips and returns string or empty."""
        if idx is None or idx < 0 or idx >= len(row):
            return ""
        v = row[idx]
        if v is None:
            return ""
        return str(v).strip()

    def _is_number_only(s: str) -> bool:
        """True if the string is purely numeric (a weight, not a KRA title)."""
        try:
            float(s.replace("%", "").replace(",", ""))
            return True
        except ValueError:
            return False

    def _parse_weight(s: str):
        """Convert weight string to integer percentage."""
        try:
            w = float(s.replace("%", "").replace(",", "").strip())
            return round(w * 100) if 0 < w <= 1.0 else round(w)
        except ValueError:
            return None

    def _is_header_row(row) -> bool:
        """True if this row looks like a column header (contains 'KRA', 'KPI', etc.)."""
        cells = [str(c).lower().strip() if c is not None else "" for c in row[:5]]
        return (
            any("kra" in c or "key result" in c for c in cells)
            and any("kpi" in c or "key performance" in c or "indicator" in c for c in cells)
        )

    # Detect header row and skip it
    start_idx = 0
    for i, r in enumerate(rows[:5]):
        if _is_header_row(r):
            start_idx = i + 1
            break

    # Determine which columns hold KRA (col_a), KPI (col_b), date (col_c), and weight (col_weight).
    # Heuristically detect columns by searching for keywords in the first 5 rows
    col_a = 0  # Default KRA title/weight
    col_b = 1  # Default KPI title
    col_c = None  # Default date
    col_weight = None  # Default weight
    
    header_found = False
    for i, row in enumerate(rows[:5]):
        cells = [str(c).lower().strip() if c is not None else "" for c in row]
        has_kra = any("kra" in c or "key result" in c or "result area" in c for c in cells)
        has_kpi = any("kpi" in c or "key performance" in c or "indicator" in c or "metric" in c for c in cells)
        if has_kra or has_kpi:
            header_found = True
            for idx, cell in enumerate(cells):
                if "kra" in cell or "key result" in cell or "result area" in cell:
                    col_a = idx
                elif "kpi" in cell or "key performance" in cell or "indicator" in cell or "metric" in cell:
                    col_b = idx
                elif "date" in cell or "target" in cell or "timeline" in cell:
                    col_c = idx
                elif "weight" in cell or "%" in cell:
                    col_weight = idx
            break

    if not header_found:
        num_cols = max((len(r) for r in rows if r), default=1)
        col_a = 0
        col_b = 1 if num_cols >= 2 else 0
        col_c = 2 if num_cols >= 3 else None

    kras = []
    current_kra: dict | None = None

    for r in rows[start_idx:]:
        # Skip fully empty rows
        if not any(c is not None and str(c).strip() for c in r):
            continue

        cell_a = _get(r, col_a)
        cell_b = _get(r, col_b)
        cell_c = _get(r, col_c)
        cell_w = _get(r, col_weight) if col_weight is not None else ""

        # Clean up date strings (openpyxl returns datetime objects as strings like "2026-01-20 00:00:00")
        if cell_c and "00:00:00" in cell_c:
            cell_c = cell_c.split(" ")[0]

        # Parse weight from either the weight column or column A (if it is purely numeric weight format)
        raw_w = cell_w if col_weight is not None and cell_w else (cell_a if _is_number_only(cell_a) else "")
        weight = _parse_weight(raw_w) if raw_w else None

        # ── Determine row type ───────────────────────────────────────────────
        # Case 1: Col A is non-empty and NOT a number only -> new KRA title
        if cell_a and not _is_number_only(cell_a):
            current_kra = {"title": cell_a, "description": "", "weight": weight, "kpis": []}
            kras.append(current_kra)
            # The KPI on the same row as the KRA title is the first KPI
            if cell_b:
                kpis = split_kpi_text(cell_b)
                for kpi_title in kpis:
                    kpi = {"title": kpi_title, "description": ""}
                    if cell_c:
                        kpi["target_date"] = cell_c
                    current_kra["kpis"].append(kpi)
            continue

        # Case 2: Col A has a number or is empty -> KPI row under the current KRA
        if not cell_a or _is_number_only(cell_a):
            if current_kra is not None and current_kra.get("weight") is None and weight is not None:
                current_kra["weight"] = weight
            
            if cell_b:
                if current_kra is None:
                    current_kra = {"title": "General", "description": "", "weight": weight, "kpis": []}
                    kras.append(current_kra)
                kpis = split_kpi_text(cell_b)
                for kpi_title in kpis:
                    kpi = {"title": kpi_title, "description": ""}
                    if cell_c:
                        kpi["target_date"] = cell_c
                    current_kra["kpis"].append(kpi)
            continue

    # Remove any KRAs that ended up with no KPIs (stray rows)
    kras = [k for k in kras if k["kpis"]]
    return kras



async def infer_jd_from_kras(employee_id: str, employee_name: str, kras: list) -> dict:
    """
    Calls Gemini to infer the employee's JD shell/profile context based on the parsed KRAs.
    Does not modify the KRAs/KPIs themselves.
    """
    from langchain_google_genai import ChatGoogleGenerativeAI
    from langchain_core.prompts import ChatPromptTemplate
    from app.core.config import settings
    import json
    
    llm = ChatGoogleGenerativeAI(
        google_api_key=settings.GEMINI_API_KEY,
        model="gemini-2.5-flash",
        temperature=0.2,
        max_output_tokens=4000,
        response_mime_type="application/json",
    )
    
    kras_summary = []
    for idx, k in enumerate(kras):
        kpis_str = ", ".join([kp["title"] for kp in k.get("kpis", [])])
        kras_summary.append(f"KRA {idx+1}: {k['title']} ({k.get('description', '')}) -> KPIs: {kpis_str}")
    kras_text = "\n".join(kras_summary)
    
    prompt = ChatPromptTemplate.from_template("""
    You are an expert HR analyst. Given an employee's Key Result Areas (KRAs) and Key Performance Indicators (KPIs), infer a professional Job Description (JD) matching this role.
    
    Employee Name: {employee_name}
    Employee ID: {employee_id}
    
    EMPLOYEE'S KRAs AND KPIs:
    {kras_text}
    
    RETURN A JSON OBJECT WITH THE FOLLOWING STRUCTURE (ONLY the "jd" block):
    {{
      "jd": {{
        "role_title": "Job title/position (e.g. Senior Software Engineer)",
        "department": "Department or function (e.g. Engineering)",
        "level": "Seniority level (Junior, Mid, Senior, Lead, Head, Director, VP)",
        "purpose": "Inferred main purpose/mission of the role (50-100 words)",
        "tasks": ["Key responsibility 1", "Key responsibility 2", ...],
        "priority_tasks": ["Top priority task 1", "Top priority task 2", ...],
        "skills": ["Required technical or soft skill 1", "Required technical or soft skill 2", ...],
        "tools": ["Tool 1", "Tool 2", ...],
        "technologies": ["Technology/language 1", "Technology/language 2", ...],
        "qualifications": {{
          "education": "Required education (e.g. Bachelor's in CS)",
          "experience_years": "Years of experience (e.g. 5+ years)",
          "certifications": []
        }},
        "working_relationships": {{
          "reports_to": "Reporting manager role title",
          "team_size": "Approximate team size (e.g. 0-5)",
          "stakeholders": []
        }}
      }}
    }}
    
    CRITICAL CONSTRAINTS:
    1. Return ONLY valid JSON matching the structure.
    2. Do NOT add any markdown formatting wrapper.
    """)
    
    chain = prompt | llm
    response = await chain.ainvoke({
        "employee_name": employee_name,
        "employee_id": employee_id,
        "kras_text": kras_text
    })
    
    raw_content = response.content.strip()
    if raw_content.startswith("```json"):
        raw_content = raw_content[7:]
    elif raw_content.startswith("```"):
        raw_content = raw_content[3:]
    if raw_content.endswith("```"):
        raw_content = raw_content[:-3]
    raw_content = raw_content.strip()
    
    return json.loads(raw_content).get("jd", {})


def _extract_excel_text(file_bytes: bytes, file_type_lower: str) -> str:
    """
    Extract text from Excel files (.xlsx or .xls) for LLM processing.
    Converts all sheets into a readable text representation.
    """
    import io
    lines = []
    
    if file_type_lower == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"=== Sheet: {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    # Skip completely empty rows
                    if not any(cell is not None for cell in row):
                        continue
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    lines.append(row_text)
                lines.append("")
            wb.close()
        except ImportError:
            raise Exception(
                "openpyxl is required to parse .xlsx files. "
                "Please install it: pip install openpyxl"
            )
    elif file_type_lower == "xls":
        try:
            import xlrd  # type: ignore
            wb = xlrd.open_workbook(file_contents=file_bytes)
            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                lines.append(f"=== Sheet: {ws.name} ===")
                for row_idx in range(ws.nrows):
                    row = ws.row_values(row_idx)
                    if not any(cell != "" and cell is not None for cell in row):
                        continue
                    row_text = "\t".join(str(cell) for cell in row)
                    lines.append(row_text)
                lines.append("")
        except ImportError:
            raise Exception(
                "xlrd is required to parse .xls files. "
                "Please install it: pip install xlrd"
            )

    return "\n".join(lines)


async def process_kra_kpi_document(
    db: AsyncSession,
    file_bytes: bytes,
    filename: str,
    file_type: str,  # "pdf", "docx", "xlsx", or "xls"
    employee_id: str,
    employee_name: str,
    admin_role: str,
) -> dict:
    from app.services.pdf_processor import PDFProcessor
    from app.services.docx_processor import DOCXProcessor
    from langchain_google_genai import ChatGoogleGenerativeAI
    from langchain_core.prompts import ChatPromptTemplate
    from app.core.config import settings
    from app.core.cache import invalidate_pattern
    from app.routers.admin_jd_routes import (
        _ensure_employee_record,
        generate_jd_text_from_structured_data,
        transform_reference_to_jd_session_schema,
    )
    import json
    import io

    file_type_lower = file_type.lower()
    
    # 1. Validate & Extract Text
    if file_type_lower == "pdf":
        is_valid, error_msg = PDFProcessor.validate_pdf(file_bytes)
        if not is_valid:
            raise Exception(f"Invalid PDF: {error_msg}")
        text = PDFProcessor.extract_text(file_bytes)
    elif file_type_lower == "docx":
        is_valid, error_msg = DOCXProcessor.validate_docx(file_bytes)
        if not is_valid:
            raise Exception(f"Invalid DOCX: {error_msg}")
        text = DOCXProcessor.extract_text(file_bytes)
    elif file_type_lower in ("xlsx", "xls"):
        try:
            kras = parse_kra_kpi_excel(file_bytes, file_type_lower)
            if kras:
                jd_session_result = await db.execute(
                    select(JDSession)
                    .where(JDSession.employee_id == employee_id)
                    .order_by(JDSession.updated_at.desc())
                )
                jd_session = jd_session_result.scalars().first()
                
                jd_data = {}
                if jd_session and jd_session.jd_structured:
                    structured = jd_session.jd_structured
                    jd_data = {
                        "role_title": jd_session.title or structured.get("role_title", "Unknown Role"),
                        "department": jd_session.department or structured.get("department", "Unknown"),
                        "level": structured.get("level", "Mid"),
                        "purpose": structured.get("purpose", ""),
                        "tasks": structured.get("tasks", []),
                        "priority_tasks": structured.get("priority_tasks", []),
                        "skills": structured.get("skills", []),
                        "tools": structured.get("tools", []),
                        "technologies": structured.get("technologies", []),
                        "qualifications": structured.get("qualifications", {}),
                        "working_relationships": structured.get("working_relationships", {})
                    }
                else:
                    try:
                        jd_data = await infer_jd_from_kras(employee_id, employee_name, kras)
                    except Exception as infer_err:
                        logger.error(f"Inferring JD failed: {infer_err}")
                        jd_data = {
                            "role_title": "Position for " + employee_name,
                            "department": "General",
                            "level": "Mid",
                            "purpose": "Role mapped from uploaded KRA/KPI framework."
                        }
                
                return {
                    "jd": jd_data,
                    "kra_kpi": {
                        "kras": split_all_kpis_in_kras(kras)
                    }
                }
        except Exception as excel_err:
            logger.error(f"Direct Excel parsing failed: {excel_err}. Falling back to text extraction + LLM.")
        
        text = _extract_excel_text(file_bytes, file_type_lower)
    else:
        raise Exception(f"Unsupported file type: {file_type}")

    if not text or len(text.strip()) == 0:
        raise Exception("Extracted text is empty — the file may be blank or unreadable")



    # 2. Query active JD session
    jd_session_result = await db.execute(
        select(JDSession)
        .where(JDSession.employee_id == employee_id)
        .order_by(JDSession.updated_at.desc())
    )
    jd_session = jd_session_result.scalars().first()

    existing_jd_text = ""
    if jd_session and jd_session.jd_text:
        existing_jd_text = jd_session.jd_text

    # 3. Call Gemini to Parse KRA/KPI and JD
    llm = ChatGoogleGenerativeAI(
        google_api_key=settings.GEMINI_API_KEY,
        model="gemini-2.5-flash",
        temperature=0.2,
        max_output_tokens=8192,
        response_mime_type="application/json",
    )
    
    context = f"Employee ID: {employee_id}\nEmployee Name: {employee_name}\n"
    if existing_jd_text:
        context += f"\nEMPLOYEE'S EXISTING JOB DESCRIPTION:\n{existing_jd_text}\n"
    
    prompt = ChatPromptTemplate.from_template("""
    You are an expert HR analyst and organizational designer.
    Extract the KRA (Key Result Area) and KPI (Key Performance Indicator) framework from the provided text.
    Also, infer or align a professional Job Description (JD) matching the role and reflecting the extracted KRA/KPI responsibilities.
    Do NOT generate or assign any weights to the KRAs.
    Return the result as a structured JSON object.

    {context}

    KRA/KPI DOCUMENT CONTENT:
    {text}

    EXTRACT AND RETURN A JSON OBJECT WITH THE FOLLOWING STRUCTURE:
    {{
      "jd": {{
        "role_title": "Job title/position (e.g. Senior Software Engineer)",
        "department": "Department or function (e.g. Engineering)",
        "level": "Seniority level (Junior, Mid, Senior, Lead, Head, Director, VP)",
        "purpose": "Inferred main purpose/mission of the role (50-100 words)",
        "tasks": ["Key responsibility 1", "Key responsibility 2", ...],
        "priority_tasks": ["Top priority task 1", "Top priority task 2", ...],
        "skills": ["Required technical or soft skill 1", "Required technical or soft skill 2", ...],
        "tools": ["Tool 1", "Tool 2", ...],
        "technologies": ["Technology/language 1", "Technology/language 2", ...],
        "qualifications": {{
          "education": "Required education (e.g. Bachelor's in CS)",
          "experience_years": "Years of experience (e.g. 5+ years)",
          "certifications": []
        }},
        "working_relationships": {{
          "reports_to": "Reporting manager role title",
          "team_size": "Approximate team size (e.g. 0-5)",
          "stakeholders": []
        }}
      }},
      "kra_kpi": {{
        "kras": [
          {{
            "kra_id": "kra_001",
            "title": "Title of the Key Result Area (e.g. Code Quality)",
            "description": "Description of the Key Result Area",
            "manager_impact": "How this KRA impacts the manager's success",
            "source_tasks": ["Key responsibility 1"],
            "kpis": [
              {{
                "kpi_id": "kpi_001",
                "title": "Title/Metric description (e.g. Sprint Delivery Rate)",
                "description": "Measurement description or targets (e.g. 90% sprint tasks delivered on time)"
              }}
            ]
          }}
        ]
      }}
    }}

    CRITICAL CONSTRAINTS:
    1. Return ONLY valid JSON. Do not include markdown code block formatting (like ```json ... ```).
    2. Do NOT generate or assign any weights for KRAs/KPIs.
    3. Make sure to generate unique IDs like kra_001, kra_002, kpi_001, kpi_002, etc.
    4. Be extremely thorough. Do not summarize or omit responsibilities or indicators.
    5. Keep descriptions clear and concise to prevent response truncation.
    """)

    # Format prompt and call
    chain = prompt | llm
    response = await chain.ainvoke({"context": context, "text": text})
    
    # Parse JSON
    raw_content = response.content.strip()
    if raw_content.startswith("```json"):
        raw_content = raw_content[7:]
    elif raw_content.startswith("```"):
        raw_content = raw_content[3:]
    if raw_content.endswith("```"):
        raw_content = raw_content[:-3]
    raw_content = raw_content.strip()
    
    try:
        parsed_data = json.loads(raw_content)
    except Exception as e:
        logger.error(f"Failed to parse LLM response: {raw_content}")
        raise Exception(f"AI parsing did not return valid JSON: {str(e)}")

    extracted_jd = parsed_data.get("jd", {})
    extracted_kra_kpi = parsed_data.get("kra_kpi", {})
    kras_list = extracted_kra_kpi.get("kras", [])
    
    if not kras_list:
        raise Exception("No KRAs or KPIs could be extracted from the document")

    extracted_kra_kpi["kras"] = split_all_kpis_in_kras(kras_list)
    parsed_data["kra_kpi"] = extracted_kra_kpi

    return parsed_data


async def analyze_kra_kpi_text(
    employee_id: str,
    employee_name: str,
    content: str,
) -> dict:
    """
    Directly call Gemini on the pasted raw text content to extract structured KRA/KPI and JD.
    """
    from langchain_google_genai import ChatGoogleGenerativeAI
    from langchain_core.prompts import ChatPromptTemplate
    from app.core.config import settings
    import json
    
    if not content or len(content.strip()) == 0:
        raise Exception("Pasted content is empty")

    llm = ChatGoogleGenerativeAI(
        google_api_key=settings.GEMINI_API_KEY,
        model="gemini-2.5-flash",
        temperature=0.2,
        max_output_tokens=8192,
        response_mime_type="application/json",
    )
    
    context = f"Employee ID: {employee_id}\nEmployee Name: {employee_name}\n"
    
    prompt = ChatPromptTemplate.from_template("""
    You are an expert HR analyst and organizational designer.
    Extract the KRA (Key Result Area) and KPI (Key Performance Indicator) framework from the provided text, and infer a professional Job Description (JD) matching the role.
    Do NOT generate or assign any weights to the KRAs.
    Return the result as a structured JSON object.

    {context}

    PASTED KRA/KPI CONTENT:
    {text}

    EXTRACT AND RETURN A JSON OBJECT WITH THE FOLLOWING STRUCTURE:
    {{
      "jd": {{
        "role_title": "Job title/position (e.g. Senior Software Engineer)",
        "department": "Department or function (e.g. Engineering)",
        "level": "Seniority level (Junior, Mid, Senior, Lead, Head, Director, VP)",
        "purpose": "Inferred main purpose/mission of the role (50-100 words)",
        "tasks": ["Key responsibility 1", "Key responsibility 2", ...],
        "priority_tasks": ["Top priority task 1", "Top priority task 2", ...],
        "skills": ["Required technical or soft skill 1", "Required technical or soft skill 2", ...],
        "tools": ["Tool 1", "Tool 2", ...],
        "technologies": ["Technology/language 1", "Technology/language 2", ...],
        "qualifications": {{
          "education": "Required education (e.g. Bachelor's in CS)",
          "experience_years": "Years of experience (e.g. 5+ years)",
          "certifications": []
        }},
        "working_relationships": {{
          "reports_to": "Reporting manager role title",
          "team_size": "Approximate team size (e.g. 0-5)",
          "stakeholders": []
        }}
      }},
      "kra_kpi": {{
        "kras": [
          {{
            "kra_id": "kra_001",
            "title": "Title of the Key Result Area (e.g. Code Quality)",
            "description": "Description of the Key Result Area",
            "manager_impact": "How this KRA impacts the manager's success",
            "source_tasks": ["Key responsibility 1"],
            "kpis": [
              {{
                "kpi_id": "kpi_001",
                "title": "Title/Metric description (e.g. Sprint Delivery Rate)",
                "description": "Measurement description or targets (e.g. 90% sprint tasks delivered on time)"
              }}
            ]
          }}
        ]
      }}
    }}

    CRITICAL CONSTRAINTS:
    1. Return ONLY valid JSON. Do not include markdown code block formatting (like ```json ... ```).
    2. Do NOT generate or assign any weights for KRAs/KPIs.
    3. Make sure to generate unique IDs like kra_001, kra_002, kpi_001, kpi_002, etc.
    4. Be extremely thorough. Do not summarize or omit responsibilities or indicators.
    5. Keep descriptions clear and concise to prevent response truncation.
    """)

    chain = prompt | llm
    response = await chain.ainvoke({"context": context, "text": content})
    
    raw_content = response.content.strip()
    if raw_content.startswith("```json"):
        raw_content = raw_content[7:]
    elif raw_content.startswith("```"):
        raw_content = raw_content[3:]
    if raw_content.endswith("```"):
        raw_content = raw_content[:-3]
    raw_content = raw_content.strip()
    
    try:
        parsed_data = json.loads(raw_content)
    except Exception as e:
        logger.error(f"Failed to parse LLM response: {raw_content}")
        raise Exception(f"AI parsing did not return valid JSON: {str(e)}")

    if "kra_kpi" in parsed_data and "kras" in parsed_data["kra_kpi"]:
        parsed_data["kra_kpi"]["kras"] = split_all_kpis_in_kras(parsed_data["kra_kpi"]["kras"])

    return parsed_data


async def save_kra_kpi_from_paste(
    db: AsyncSession,
    employee_id: str,
    employee_name: str,
    jd_data: dict,
    kra_kpi_data: dict,
    admin_role: str,
) -> dict:
    """
    Save the confirmed JD and admin-uploaded KRA/KPI to the separate uploaded_kra_kpis table.
    """
    import uuid
    from datetime import datetime, timezone
    from app.core.cache import invalidate_pattern
    from app.routers.admin_jd_routes import (
        _ensure_employee_record,
        generate_jd_text_from_structured_data,
        transform_reference_to_jd_session_schema,
    )
    from app.models.kra_kpi_model import UploadedKRAKPI
    
    # 1. Ensure Employee Record
    employee = await _ensure_employee_record(db, employee_id, employee_name, jd_data.get("department"))

    # 2. Check/Create JD Session
    # Retrieve active JD session
    jd_session_result = await db.execute(
        select(JDSession)
        .where(JDSession.employee_id == employee_id)
        .order_by(JDSession.updated_at.desc())
    )
    jd_session = jd_session_result.scalars().first()
    
    # Transform structured JD data
    transformed_jd_structured = transform_reference_to_jd_session_schema(jd_data)
    jd_text = generate_jd_text_from_structured_data(transformed_jd_structured)

    if not jd_session:
        jd_session = JDSession(
            id=uuid.uuid4(),
            employee_id=employee_id,
            title=jd_data.get("role_title", "Unknown Role"),
            department=jd_data.get("department", "Unknown"),
            jd_text=jd_text,
            jd_structured=transformed_jd_structured,
            status="approved",
            version=1
        )
        db.add(jd_session)
        await db.flush()
    else:
        # Update existing session to approved
        jd_session.title = jd_data.get("role_title", "Unknown Role")
        jd_session.department = jd_data.get("department", "Unknown")
        jd_session.jd_text = jd_text
        jd_session.jd_structured = transformed_jd_structured
        jd_session.status = "approved"
        await db.flush()

    # 3. Check/Create Uploaded KRA/KPI
    uploaded_result = await db.execute(
        select(UploadedKRAKPI).where(UploadedKRAKPI.employee_id == employee_id)
    )
    uploaded_record = uploaded_result.scalars().first()

    # Clean the kras list to only keep title, description, and kpis (with title and description)
    kras_clean = []
    # Make sure we split any remaining concatenated KPIs before clean-up!
    kras_to_save = split_all_kpis_in_kras(kra_kpi_data.get("kras", []))
    for kra in kras_to_save:
        kpis_clean = []
        for kpi in kra.get("kpis", []):
            kpis_clean.append({
                "title": kpi.get("title", ""),
                "description": kpi.get("description", "")
            })
        kras_clean.append({
            "title": kra.get("title", ""),
            "description": kra.get("description", ""),
            "kpis": kpis_clean
        })

    now = datetime.now(timezone.utc)
    if not uploaded_record:
        uploaded_record = UploadedKRAKPI(
            id=uuid.uuid4(),
            employee_id=employee_id,
            employee_name=employee_name,
            kras={"kras": kras_clean},
            created_at=now,
            updated_at=now
        )
        db.add(uploaded_record)
    else:
        uploaded_record.employee_name = employee_name
        uploaded_record.kras = {"kras": kras_clean}
        uploaded_record.updated_at = now

    await db.commit()
    await db.refresh(jd_session)
    await db.refresh(uploaded_record)

    # Invalidate caches
    await invalidate_pattern(f"jds:employee:{employee_id}")
    await invalidate_pattern(f"cache:jd_detail:*{jd_session.id}*")
    
    return {
        "jd_session_id": str(jd_session.id),
        "uploaded_kra_kpi_id": str(uploaded_record.id),
        "employee_id": employee_id,
        "employee_name": employee_name,
        "role_title": jd_data.get("role_title"),
        "department": jd_data.get("department"),
        "kras_count": len(kras_clean),
        "status": "success",
    }


async def sync_kra_kpi_session_to_db(
    db: AsyncSession,
    session_id: str,
    conversation_state: dict,
    conversation_history: list,
) -> KRAKPISession:
    import uuid
    from sqlalchemy import delete
    from sqlalchemy.orm.attributes import flag_modified
    from app.models.kra_kpi_model import KRAKPIConversationTurn

    session_uuid = uuid.UUID(session_id) if isinstance(session_id, str) else session_id
    result = await db.execute(select(KRAKPISession).where(KRAKPISession.id == session_uuid))
    record = result.scalar_one_or_none()

    if record:
        record.conversation_state = conversation_state
        flag_modified(record, "conversation_state")

        # Sync conversation history turns
        await db.execute(
            delete(KRAKPIConversationTurn).where(KRAKPIConversationTurn.session_id == session_uuid)
        )
        for idx, turn in enumerate(conversation_history):
            new_turn = KRAKPIConversationTurn(
                session_id=session_uuid,
                turn_index=idx,
                role=turn.get("role", "unknown"),
                content=turn.get("content", ""),
            )
            db.add(new_turn)

        await db.commit()
        await db.refresh(record, ["conversation_turns"])
    return record

