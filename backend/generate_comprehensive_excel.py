import asyncio
import os
import pandas as pd
from sqlalchemy import text
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.database import engine

async def generate_report():
    print("Connecting to database and fetching dataset...")
    async with engine.connect() as conn:
        # 1. Shared approved JD map
        res_approved = await conn.execute(text("""
            SELECT id, department, title
            FROM jd_sessions
            WHERE status = 'approved'
              AND title IS NOT NULL
              AND department IS NOT NULL
        """))
        approved_map = {}
        for row in res_approved.fetchall():
            dept_key = row.department.strip().lower() if row.department else ""
            title_key = row.title.strip().lower() if row.title else ""
            approved_map[(dept_key, title_key)] = str(row.id)

        # 2. Latest JD Session per employee
        res_jd = await conn.execute(text("""
            WITH RankedJDs AS (
                SELECT 
                    id as jd_id, employee_id, status as jd_status, updated_at,
                    ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
                FROM jd_sessions
            )
            SELECT jd_id, employee_id, jd_status, updated_at
            FROM RankedJDs WHERE rn = 1
        """))
        jd_map = {row.employee_id: dict(row._mapping) for row in res_jd.fetchall()}

        # 3. Latest KRA KPI Session per employee
        res_kra = await conn.execute(text("""
            WITH RankedKRAs AS (
                SELECT 
                    id as kra_id, employee_id, status as kra_status, generation_step, updated_at,
                    ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
                FROM kra_kpi_sessions
            )
            SELECT kra_id, employee_id, kra_status, generation_step, updated_at
            FROM RankedKRAs WHERE rn = 1
        """))
        kra_map = {row.employee_id: dict(row._mapping) for row in res_kra.fetchall()}

        # 4. Uploaded KRAs map
        res_up = await conn.execute(text('SELECT employee_id, updated_at FROM uploaded_kra_kpis'))
        uploaded_map = {row.employee_id: dict(row._mapping) for row in res_up.fetchall()}

        # 5. Organogram employees
        res_org = await conn.execute(text("""
            SELECT 
                code, employee_name, designation, department, date_of_joining,
                location, reporting_manager_code, reporting_manager, joblevel
            FROM organogram
            ORDER BY employee_name ASC
        """))
        organogram_employees = [dict(r._mapping) for r in res_org.fetchall()]

    # 6. Read Roorkee Excel file if present
    roorkee_file = "../Roorkee__PP_Dataset_JD_s_Link_s_16-07-2026 -Updated.xlsx"
    df_roorkee_raw = pd.DataFrame()
    if os.path.exists(roorkee_file):
        try:
            df_roorkee_raw = pd.read_excel(roorkee_file)
        except Exception as e:
            print(f"Warning: Could not read {roorkee_file}: {e}")
    
    # Process Organogram employees
    processed_records = []
    seen_codes = set()

    for emp in organogram_employees:
        emp_code = str(emp["code"]).strip() if emp["code"] else ""
        seen_codes.add(emp_code)
        
        emp_name = emp["employee_name"] or ""
        designation = emp["designation"] or ""
        department = emp["department"] or ""
        doj = emp["date_of_joining"] or ""
        location = emp["location"] or ""
        manager_code = emp["reporting_manager_code"] or ""
        manager_name = emp["reporting_manager"] or ""
        job_level = emp["joblevel"] or ""

        # Retrieve JD Info
        jd_info = jd_map.get(emp_code)
        raw_jd_status = jd_info["jd_status"] if jd_info else None

        # Retrieve KRA Info
        kra_info = kra_map.get(emp_code)
        raw_kra_status = kra_info["kra_status"] if kra_info else None
        is_uploaded_kra = emp_code in uploaded_map

        # Check shared template availability
        dept_key = department.strip().lower() if department else ""
        desig_key = designation.strip().lower() if designation else ""
        has_shared_approved_jd = (dept_key, desig_key) in approved_map

        # Map JD Status
        if raw_jd_status == "approved":
            jd_status_label = "Approved"
            jd_details = "Approved (Personal JD)"
            jd_is_complete = True
        elif raw_jd_status == "sent_to_hr":
            jd_status_label = "Under Review (HR)"
            jd_details = "Pending HR Approval"
            jd_is_complete = False
        elif raw_jd_status == "sent_to_manager":
            jd_status_label = "Under Review (Manager)"
            jd_details = f"Pending Manager Approval ({manager_name})"
            jd_is_complete = False
        elif raw_jd_status == "collecting":
            if has_shared_approved_jd:
                jd_status_label = "Approved (Shared)"
                jd_details = "Available (Shared Role Approved)"
                jd_is_complete = True
            else:
                jd_status_label = "In Progress"
                jd_details = "Collecting Data / Draft"
                jd_is_complete = False
        elif raw_jd_status == "jd_generated":
            if has_shared_approved_jd:
                jd_status_label = "Approved (Shared)"
                jd_details = "Available (Shared Role Approved)"
                jd_is_complete = True
            else:
                jd_status_label = "In Progress"
                jd_details = "JD Draft Generated"
                jd_is_complete = False
        elif raw_jd_status == "manager_rejected":
            jd_status_label = "In Progress (Revision)"
            jd_details = f"Rejected by Manager ({manager_name})"
            jd_is_complete = False
        elif raw_jd_status == "hr_rejected":
            jd_status_label = "In Progress (Revision)"
            jd_details = "Rejected by HR"
            jd_is_complete = False
        elif raw_jd_status == "draft":
            if has_shared_approved_jd:
                jd_status_label = "Approved (Shared)"
                jd_details = "Available (Shared Role Approved)"
                jd_is_complete = True
            else:
                jd_status_label = "In Progress"
                jd_details = "Draft"
                jd_is_complete = False
        else:
            if has_shared_approved_jd:
                jd_status_label = "Approved (Shared)"
                jd_details = "Available (Shared Role Approved)"
                jd_is_complete = True
            else:
                jd_status_label = "Not Started"
                jd_details = "No JD Session Created"
                jd_is_complete = False

        # Map KRA/KPI Status
        if is_uploaded_kra:
            kra_status_label = "Uploaded / Active"
            kra_details = "Admin Uploaded Direct KRA/KPI"
            kra_is_complete = True
        elif raw_kra_status == "confirmed":
            kra_status_label = "Confirmed / Active"
            kra_details = "KRA & KPI Confirmed by Employee"
            kra_is_complete = True
        elif raw_kra_status == "sent_to_hr":
            kra_status_label = "Under Review (HR)"
            kra_details = "Pending HR Approval"
            kra_is_complete = False
        elif raw_kra_status == "sent_to_manager":
            kra_status_label = "Under Review (Manager)"
            kra_details = f"Pending Manager Approval ({manager_name})"
            kra_is_complete = False
        elif raw_kra_status == "draft":
            kra_status_label = "In Progress"
            kra_details = "KRA Draft In Progress"
            kra_is_complete = False
        else:
            kra_status_label = "Not Started"
            kra_details = "No KRA Session Created"
            kra_is_complete = False

        # Map Manager Approval Flow & Action Required
        if raw_jd_status == "sent_to_manager" or raw_kra_status == "sent_to_manager":
            approval_flow_status = "Under Review - Pending Manager"
            approval_action_required = f"Pending Manager Approval: {manager_name} ({manager_code})"
            overall_status = "Under Review"
        elif raw_jd_status == "sent_to_hr" or raw_kra_status == "sent_to_hr":
            approval_flow_status = "Under Review - Pending HR"
            approval_action_required = "Pending HR Approval"
            overall_status = "Under Review"
        elif jd_is_complete and kra_is_complete:
            approval_flow_status = "Completed & Active"
            approval_action_required = "Fully Approved & Working"
            overall_status = "Working / Completed"
        elif jd_is_complete:
            approval_flow_status = "JD Approved - KRA Pending"
            approval_action_required = "Employee needs to complete KRA/KPI setup"
            overall_status = "Working (JD Active)"
        elif raw_jd_status in ["collecting", "jd_generated", "draft", "manager_rejected", "hr_rejected"] or raw_kra_status == "draft":
            approval_flow_status = "In Progress - Employee Draft"
            approval_action_required = "Employee working on draft/revisions"
            overall_status = "In Progress"
        else:
            approval_flow_status = "Not Started"
            approval_action_required = "Employee needs to initiate JD creation"
            overall_status = "Not Started"

        # Determine Primary Department Group (Separating PNS Plant and PP-Roorkee)
        dept_lower = department.lower()
        loc_lower = location.lower()

        is_roorkee_excel = False
        if not df_roorkee_raw.empty and "Code" in df_roorkee_raw.columns:
            is_roorkee_excel = emp_code in df_roorkee_raw["Code"].astype(str).tolist()

        if loc_lower == "r&d" or "r&d" in dept_lower or "research" in dept_lower:
            dept_group = "R&D Corporate"
        elif is_roorkee_excel or "roorkee" in loc_lower or "roorkee" in dept_lower or (loc_lower == "factory" and any(q in dept_lower for q in ["qa", "qc", "quality", "regulatory", "hrd (plant)"])):
            dept_group = "PP-Roorkee"
        elif loc_lower == "factory" or any(p in dept_lower for p in ["production", "maintenance", "stores", "plant", "warehouse", "engineering", "pns"]):
            dept_group = "PNS Plant"
        else:
            dept_group = "Corporate & Head Office"

        sso_link = f"https://jd.pulsepharma.net/sso?employee_id={emp_code}"

        processed_records.append({
            "Employee Code": emp_code,
            "Employee Name": emp_name,
            "Designation": designation,
            "Department": department,
            "Job Level": job_level,
            "Location": location,
            "Department Group": dept_group,
            "Reporting Manager Code": manager_code,
            "Reporting Manager Name": manager_name,
            "JD Status": jd_status_label,
            "JD Details": jd_details,
            "JD Completed": "Yes" if jd_is_complete else "No",
            "KRA & KPI Status": kra_status_label,
            "KRA Details": kra_details,
            "KRA Completed": "Yes" if kra_is_complete else "No",
            "Overall Status": overall_status,
            "Approval Flow Stage": approval_flow_status,
            "Action Required / Approver": approval_action_required,
            "Date of Joining": doj,
            "SSO Portal Link": sso_link
        })

    # Add Roorkee employees from Excel if not present in organogram
    if not df_roorkee_raw.empty and "Code" in df_roorkee_raw.columns:
        for _, r_row in df_roorkee_raw.iterrows():
            rk_code = str(r_row.get("Code", "")).strip()
            if rk_code and rk_code not in seen_codes:
                seen_codes.add(rk_code)
                emp_name = str(r_row.get("Employee_Name", r_row.get("Emp Name", "")))
                designation = str(r_row.get("Designation", ""))
                department = str(r_row.get("Department", ""))
                job_level = str(r_row.get("Joblevel", r_row.get("Job_level", "")))
                location = str(r_row.get("Location", "Roorkee"))
                manager_code = str(r_row.get("Reporting_Manager_Code", r_row.get("Rep Manager Code", "")))
                manager_name = str(r_row.get("Reporting_Manager", r_row.get("Rep Manager", "")))
                doj = str(r_row.get("Date_of_joining", r_row.get("DOJ", "")))

                # Check DB maps
                jd_info = jd_map.get(rk_code)
                raw_jd_status = jd_info["jd_status"] if jd_info else None
                kra_info = kra_map.get(rk_code)
                raw_kra_status = kra_info["kra_status"] if kra_info else None
                is_uploaded_kra = rk_code in uploaded_map

                jd_status_label = "Not Started"
                jd_details = "No JD Session Created"
                jd_is_complete = False
                if raw_jd_status == "approved":
                    jd_status_label = "Approved"
                    jd_details = "Approved (Personal JD)"
                    jd_is_complete = True
                elif raw_jd_status == "sent_to_hr":
                    jd_status_label = "Under Review (HR)"
                    jd_details = "Pending HR Approval"
                elif raw_jd_status == "sent_to_manager":
                    jd_status_label = "Under Review (Manager)"
                    jd_details = f"Pending Manager Approval ({manager_name})"

                kra_status_label = "Uploaded / Active" if is_uploaded_kra else "Not Started"
                kra_details = "Admin Uploaded Direct KRA/KPI" if is_uploaded_kra else "No KRA Session Created"
                kra_is_complete = True if is_uploaded_kra else False

                if raw_jd_status == "sent_to_manager" or raw_kra_status == "sent_to_manager":
                    approval_flow_status = "Under Review - Pending Manager"
                    approval_action_required = f"Pending Manager Approval: {manager_name}"
                    overall_status = "Under Review"
                elif raw_jd_status == "sent_to_hr" or raw_kra_status == "sent_to_hr":
                    approval_flow_status = "Under Review - Pending HR"
                    approval_action_required = "Pending HR Approval"
                    overall_status = "Under Review"
                elif jd_is_complete:
                    approval_flow_status = "Completed & Active"
                    approval_action_required = "Fully Approved & Working"
                    overall_status = "Working / Completed"
                else:
                    approval_flow_status = "Not Started"
                    approval_action_required = "Employee needs to initiate JD creation"
                    overall_status = "Not Started"

                sso_link = f"https://jd.pulsepharma.net/sso?employee_id={rk_code}"

                processed_records.append({
                    "Employee Code": rk_code,
                    "Employee Name": emp_name,
                    "Designation": designation,
                    "Department": department,
                    "Job Level": job_level,
                    "Location": location,
                    "Department Group": "PP-Roorkee",
                    "Reporting Manager Code": manager_code,
                    "Reporting Manager Name": manager_name,
                    "JD Status": jd_status_label,
                    "JD Details": jd_details,
                    "JD Completed": "Yes" if jd_is_complete else "No",
                    "KRA & KPI Status": kra_status_label,
                    "KRA Details": kra_details,
                    "KRA Completed": "Yes" if kra_is_complete else "No",
                    "Overall Status": overall_status,
                    "Approval Flow Stage": approval_flow_status,
                    "Action Required / Approver": approval_action_required,
                    "Date of Joining": doj,
                    "SSO Portal Link": sso_link
                })

    df_master = pd.DataFrame(processed_records)
    print(f"Total processed employee records: {len(df_master)}")
    print("Department Group Counts:\n", df_master["Department Group"].value_counts())
    print("Overall Status Counts:\n", df_master["Overall Status"].value_counts())

    # Build individual filtered DataFrames for distinct worksheets
    df_pns = df_master[df_master["Department Group"] == "PNS Plant"].copy()
    df_roorkee = df_master[df_master["Department Group"] == "PP-Roorkee"].copy()
    df_rd = df_master[df_master["Department Group"] == "R&D Corporate"].copy()
    df_corp = df_master[df_master["Department Group"] == "Corporate & Head Office"].copy()

    # Build Department Summary Matrix
    dept_summary_records = []
    for (dept_group, dept_name), group_df in df_master.groupby(["Department Group", "Department"]):
        total_emp = len(group_df)
        jd_completed = len(group_df[group_df["JD Completed"] == "Yes"])
        jd_pending = total_emp - jd_completed
        kra_completed = len(group_df[group_df["KRA Completed"] == "Yes"])
        kra_pending = total_emp - kra_completed
        fully_completed = len(group_df[group_df["Overall Status"] == "Working / Completed"])
        pending_manager = len(group_df[group_df["Approval Flow Stage"] == "Under Review - Pending Manager"])
        pending_hr = len(group_df[group_df["Approval Flow Stage"] == "Under Review - Pending HR"])
        not_started = len(group_df[group_df["Overall Status"] == "Not Started"])
        in_progress = len(group_df[group_df["Overall Status"] == "In Progress"])
        completion_pct = round((jd_completed / total_emp) * 100, 1) if total_emp > 0 else 0.0

        dept_summary_records.append({
            "Department Group": dept_group,
            "Department Name": dept_name,
            "Total Employees": total_emp,
            "JD Completed": jd_completed,
            "JD Pending": jd_pending,
            "KRA/KPI Active": kra_completed,
            "KRA/KPI Pending": kra_pending,
            "Fully Working / Completed": fully_completed,
            "Pending Manager Approval": pending_manager,
            "Pending HR Approval": pending_hr,
            "In Draft / Revision": in_progress,
            "Not Started": not_started,
            "JD Completion %": f"{completion_pct}%"
        })

    df_dept_summary = pd.DataFrame(dept_summary_records).sort_values(by=["Department Group", "Total Employees"], ascending=[True, False])

    # Build Group Summary Matrix
    group_summary_records = []
    for dept_group, group_df in df_master.groupby("Department Group"):
        total_emp = len(group_df)
        jd_completed = len(group_df[group_df["JD Completed"] == "Yes"])
        jd_pending = total_emp - jd_completed
        kra_completed = len(group_df[group_df["KRA Completed"] == "Yes"])
        kra_pending = total_emp - kra_completed
        fully_completed = len(group_df[group_df["Overall Status"] == "Working / Completed"])
        pending_manager = len(group_df[group_df["Approval Flow Stage"] == "Under Review - Pending Manager"])
        pending_hr = len(group_df[group_df["Approval Flow Stage"] == "Under Review - Pending HR"])
        not_started = len(group_df[group_df["Overall Status"] == "Not Started"])
        completion_pct = round((jd_completed / total_emp) * 100, 1) if total_emp > 0 else 0.0

        group_summary_records.append({
            "Business Group / Unit": dept_group,
            "Total Employees": total_emp,
            "JD Completed": jd_completed,
            "JD Pending": jd_pending,
            "KRA/KPI Active": kra_completed,
            "KRA/KPI Pending": kra_pending,
            "Fully Completed": fully_completed,
            "Pending Manager Review": pending_manager,
            "Pending HR Review": pending_hr,
            "Not Started": not_started,
            "JD Completion Rate": f"{completion_pct}%"
        })
    df_group_summary = pd.DataFrame(group_summary_records)

    # Manager Pending Approvals Summary
    df_pending_mgr = df_master[df_master["Approval Flow Stage"] == "Under Review - Pending Manager"].copy()
    mgr_summary_records = []
    for (mgr_code, mgr_name), group_df in df_pending_mgr.groupby(["Reporting Manager Code", "Reporting Manager Name"]):
        dept = group_df["Department"].iloc[0] if not group_df.empty else ""
        dept_grp = group_df["Department Group"].iloc[0] if not group_df.empty else ""
        mgr_summary_records.append({
            "Reporting Manager Code": mgr_code,
            "Reporting Manager Name": mgr_name,
            "Department Group": dept_grp,
            "Department": dept,
            "Pending Employees Count": len(group_df),
            "Pending Employee Names": ", ".join(group_df["Employee Name"].tolist()[:5]) + ("..." if len(group_df) > 5 else "")
        })
    df_mgr_summary = pd.DataFrame(mgr_summary_records).sort_values(by="Pending Employees Count", ascending=False)

    # Output file paths
    output_filename = "Employee_JD_KRA_KPI_Status_Report.xlsx"
    workspace_path = f"/Users/manideekshith/Developer/JD-Agent/{output_filename}"
    artifact_path = f"/Users/manideekshith/.gemini/antigravity-cli/brain/08abd6a5-d54f-4088-91da-14564dbe10a4/{output_filename}"

    for p in [workspace_path, artifact_path]:
        os.makedirs(os.path.dirname(p), exist_ok=True)
        with pd.ExcelWriter(p, engine='openpyxl') as writer:
            # 1. Executive Summary Sheet
            df_group_summary.to_excel(writer, sheet_name='Executive Summary', startrow=2, index=False)
            
            wb = writer.book
            ws_exec = wb['Executive Summary']

            # Titles
            ws_exec.cell(row=1, column=1, value="BUSINESS GROUP COMPLETION OVERVIEW").font = Font(name='Segoe UI', size=12, bold=True, color='1F497D')
            
            start_dept_row = len(df_group_summary) + 5
            ws_exec.cell(row=start_dept_row, column=1, value="DEPARTMENT-WISE COMPLETION BREAKDOWN").font = Font(name='Segoe UI', size=12, bold=True, color='1F497D')
            
            for c_idx, col_name in enumerate(df_dept_summary.columns, 1):
                ws_exec.cell(row=start_dept_row + 1, column=c_idx, value=col_name)
            for r_idx, row in enumerate(df_dept_summary.values, start_dept_row + 2):
                for c_idx, val in enumerate(row, 1):
                    ws_exec.cell(row=r_idx, column=c_idx, value=val)

            start_mgr_row = start_dept_row + len(df_dept_summary) + 4
            ws_exec.cell(row=start_mgr_row, column=1, value="TEAM LEADS / MANAGERS PENDING ACTION OVERVIEW").font = Font(name='Segoe UI', size=12, bold=True, color='1F497D')
            
            for c_idx, col_name in enumerate(df_mgr_summary.columns, 1):
                ws_exec.cell(row=start_mgr_row + 1, column=c_idx, value=col_name)
            for r_idx, row in enumerate(df_mgr_summary.values, start_mgr_row + 2):
                for c_idx, val in enumerate(row, 1):
                    ws_exec.cell(row=r_idx, column=c_idx, value=val)

            # 2. Master & Separate Group Sheets
            df_master.to_excel(writer, sheet_name='All Employees Master', index=False)
            df_pns.to_excel(writer, sheet_name='PNS Plant', index=False)
            df_roorkee.to_excel(writer, sheet_name='PP-Roorkee', index=False)
            df_rd.to_excel(writer, sheet_name='R&D Corporate', index=False)
            df_corp.to_excel(writer, sheet_name='Corporate & Head Office', index=False)
            df_pending_mgr.to_excel(writer, sheet_name='Manager Pending Cases', index=False)

            # Styling across all worksheets
            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid') # Corporate Navy
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            even_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

            # Status fills
            fill_working = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Soft Green
            font_working = Font(name='Segoe UI', size=10, bold=True, color='276A3C')

            fill_review = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # Soft Red/Orange
            font_review = Font(name='Segoe UI', size=10, bold=True, color='C65911')

            fill_progress = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # Soft Yellow
            font_progress = Font(name='Segoe UI', size=10, bold=True, color='8A6D3B')

            fill_notstarted = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') # Light Gray
            font_notstarted = Font(name='Segoe UI', size=10, color='595959')

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.views.sheetView[0].showGridLines = True

                if sheet_name != 'Executive Summary':
                    # Format standard data sheets
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_border
                    ws.row_dimensions[1].height = 28

                    for row_idx in range(2, ws.max_row + 1):
                        ws.row_dimensions[row_idx].height = 22
                        row_fill = even_fill if row_idx % 2 == 0 else white_fill

                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.border = thin_border
                            cell.font = Font(name='Segoe UI', size=10)
                            cell.alignment = Alignment(vertical='center')
                            cell.fill = row_fill

                            val_str = str(cell.value or "")
                            
                            # Conditional styling for status values
                            if val_str in ["Approved", "Confirmed / Active", "Uploaded / Active", "Working / Completed", "Completed & Active", "Yes"]:
                                cell.fill = fill_working
                                cell.font = font_working
                            elif "Under Review" in val_str or "Pending Manager" in val_str or "Pending HR" in val_str:
                                cell.fill = fill_review
                                cell.font = font_review
                            elif "In Progress" in val_str or "Draft" in val_str:
                                cell.fill = fill_progress
                                cell.font = font_progress
                            elif "Not Started" in val_str or val_str == "No":
                                if col_idx in [10, 12, 13, 15, 16, 17]:
                                    cell.fill = fill_notstarted
                                    cell.font = font_notstarted

                            if val_str.startswith("http"):
                                cell.hyperlink = val_str
                                cell.font = Font(name='Segoe UI', size=10, color='0066CC', underline='single')

                else:
                    # Executive summary sheet formatting
                    for row_idx in range(1, ws.max_row + 1):
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                cell.font = Font(name='Segoe UI', size=10)
                                cell.border = thin_border
                                cell.alignment = Alignment(vertical='center')

                    # Format header rows in Executive summary
                    header_rows = [3, start_dept_row + 1, start_mgr_row + 1]
                    for hr in header_rows:
                        ws.row_dimensions[hr].height = 26
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=hr, column=col_idx)
                            if cell.value:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_align

                # Auto-adjust column widths
                for col in ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    max_len = 0
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

        print(f"Successfully generated: {p}")

if __name__ == "__main__":
    asyncio.run(generate_report())
