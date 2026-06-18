import asyncio
import os
import pandas as pd
from sqlalchemy import text
from app.core.database import engine

async def generate_excel():
    print("Connecting to the database...")
    async with engine.connect() as conn:
        # 1. Fetch approved templates map for department(s) and title(s)
        # These are used for "zero-bloat shared role approved JDs"
        approved_query = text("""
            SELECT id, department, title
            FROM jd_sessions
            WHERE status = 'approved'
              AND title IS NOT NULL
              AND department IS NOT NULL
        """)
        approved_res = await conn.execute(approved_query)
        approved_map = {}
        for row in approved_res.fetchall():
            dept_key = row.department.strip().lower() if row.department else ""
            title_key = row.title.strip().lower() if row.title else ""
            approved_map[(dept_key, title_key)] = str(row.id)
            
        print(f"Loaded {len(approved_map)} approved shared templates.")

        # 2. Get latest JD session status for each employee
        latest_jds_query = text("""
            WITH LatestJDs AS (
                SELECT 
                    id as jd_id,
                    employee_id, 
                    status,
                    updated_at,
                    ROW_NUMBER() OVER(PARTITION BY employee_id ORDER BY updated_at DESC) as rn
                FROM jd_sessions
            )
            SELECT employee_id, jd_id, status, updated_at
            FROM LatestJDs
            WHERE rn = 1
        """)
        latest_jds_res = await conn.execute(latest_jds_query)
        latest_jds_map = {row.employee_id: {
            "jd_id": str(row.jd_id),
            "status": row.status,
            "updated_at": row.updated_at
        } for row in latest_jds_res.fetchall()}
        
        print(f"Loaded latest JD sessions for {len(latest_jds_map)} employees.")

        # 3. Get all employees from organogram
        organogram_query = text("""
            SELECT 
                code,
                employee_name,
                designation,
                department,
                reporting_manager
            FROM organogram
            ORDER BY employee_name ASC
        """)
        organogram_res = await conn.execute(organogram_query)
        employees = [dict(r._mapping) for r in organogram_res.fetchall()]
        print(f"Retrieved {len(employees)} employees from organogram.")

        # 4. Map statuses and build the dataset
        records = []
        for i, emp in enumerate(employees):
            emp_code = emp["code"]
            emp_name = emp["employee_name"]
            designation = emp["designation"]
            department = emp["department"]
            reporting_manager = emp["reporting_manager"]
            
            # Retrieve latest JD session if any
            jd_info = latest_jds_map.get(emp_code)
            raw_status = jd_info["status"] if jd_info else None
            
            # Map status
            mapped_status = "Not Started"
            details = ""
            
            # Check for shared approved template
            dept_key = department.strip().lower() if department else ""
            desig_key = designation.strip().lower() if designation else ""
            has_shared_approved = (dept_key, desig_key) in approved_map
            
            if raw_status == "approved":
                mapped_status = "Available"
                details = "Approved (Personal JD)"
            elif raw_status == "sent_to_hr":
                mapped_status = "In Progress"
                details = "Pending HR Approval"
            elif raw_status == "sent_to_manager":
                mapped_status = "In Progress"
                details = "Pending Manager Approval"
            elif raw_status == "collecting":
                if has_shared_approved:
                    mapped_status = "Available"
                    details = "Available (Shared Role Approved)"
                else:
                    mapped_status = "In Progress"
                    details = "Collecting Data / Draft"
            elif raw_status == "jd_generated":
                if has_shared_approved:
                    mapped_status = "Available"
                    details = "Available (Shared Role Approved)"
                else:
                    mapped_status = "In Progress"
                    details = "JD Draft Generated"
            elif raw_status == "manager_rejected":
                mapped_status = "In Progress"
                details = "Rejected by Manager (Needs Revision)"
            elif raw_status == "hr_rejected":
                mapped_status = "In Progress"
                details = "Rejected by HR (Needs Revision)"
            elif raw_status == "draft":
                if has_shared_approved:
                    mapped_status = "Available"
                    details = "Available (Shared Role Approved)"
                else:
                    mapped_status = "In Progress"
                    details = "Draft"
            else:
                # No session
                if has_shared_approved:
                    mapped_status = "Available"
                    details = "Available (Shared Role Approved)"
                else:
                    mapped_status = "Not Started"
                    details = "No JD Session Created"
            
            records.append({
                "sl.no": i + 1,
                "Employee_code": emp_code,
                "employee name": emp_name,
                "designation": designation,
                "department": department,
                "jd status": mapped_status,
                "status details": details,
                "reporting manager": reporting_manager
            })
            
        # 5. Write to Excel
        df = pd.DataFrame(records)
        
        # Path details
        output_filename = "Employee_JD_Status_Report.xlsx"
        
        # Let's save in the workspace root: /Users/manideekshith/Desktop/JD-Agent
        workspace_path = f"/Users/manideekshith/Desktop/JD-Agent/{output_filename}"
        
        # And in the artifacts folder: /Users/manideekshith/.gemini/antigravity-cli/brain/829fafb2-12cc-4fb5-8496-0d1d1ed55567
        artifact_dir = "/Users/manideekshith/.gemini/antigravity-cli/brain/829fafb2-12cc-4fb5-8496-0d1d1ed55567"
        artifact_path = os.path.join(artifact_dir, output_filename)
        
        # Save to both paths
        for p in [workspace_path, artifact_path]:
            # Create directories if they do not exist
            os.makedirs(os.path.dirname(p), exist_ok=True)
            
            # We want a nicely styled excel sheet
            with pd.ExcelWriter(p, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='JD Status Report')
                
                # Auto-adjust columns width and style header
                workbook = writer.book
                worksheet = writer.sheets['JD Status Report']
                
                # Format headers
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid') # Sleek dark blue
                header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                thin_border = Border(
                    left=Side(style='thin', color='D3D3D3'),
                    right=Side(style='thin', color='D3D3D3'),
                    top=Side(style='thin', color='D3D3D3'),
                    bottom=Side(style='thin', color='D3D3D3')
                )
                
                # Format headers
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border
                
                # Zebra striping for rows
                even_fill = PatternFill(start_color='F2F5F9', end_color='F2F5F9', fill_type='solid')
                white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                for row_num in range(2, len(records) + 2):
                    fill_to_use = even_fill if row_num % 2 == 0 else white_fill
                    # Colors for status
                    status_val = worksheet.cell(row=row_num, column=6).value # 6th column is jd status
                    
                    status_fill = None
                    status_font = None
                    if status_val == "Available":
                        status_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Soft green
                        status_font = Font(name='Arial', size=10, bold=True, color='375623')
                    elif status_val == "In Progress":
                        status_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # Soft yellow
                        status_font = Font(name='Arial', size=10, bold=True, color='7F6000')
                    elif status_val == "Not Started":
                        status_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid') # Soft red/orange
                        status_font = Font(name='Arial', size=10, color='C65911')
                    
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.font = Font(name='Arial', size=10)
                        
                        # Apply alignment
                        if col_num in [1, 2, 6]: # Sl No, Code, Status
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            
                        # Apply row striping/fill
                        if col_num == 6 and status_fill:
                            cell.fill = status_fill
                            cell.font = status_font
                        else:
                            cell.fill = fill_to_use
                            
                # Auto-adjust column widths
                for col in worksheet.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                    
            print(f"Successfully saved to {p}")
            
if __name__ == "__main__":
    asyncio.run(generate_excel())
